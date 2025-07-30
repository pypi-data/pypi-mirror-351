from openpyxl import load_workbook

from ..model import LoadPageContents, PageLoader


class XlsxLoader(PageLoader):
    def load(self, path: str) -> list[LoadPageContents]:
        page_contents_list = []
        workbook = load_workbook(filename=path)
        for index, sheet in enumerate(workbook.sheetnames):
            worksheet = workbook[sheet]
            table = [[cell.value for cell in row] for row in worksheet.iter_rows()]
            contents = f"{sheet}\n" + "\n".join([",".join([str(cell) for cell in row]) for row in table])
            page_contents = LoadPageContents(
                contents=contents,
                page_number=index,
                image=None,
                tables=[table],  # type: ignore
            )
            page_contents_list.append(page_contents)
        return page_contents_list
