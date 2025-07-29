# encoding: utf-8
# desc: 
# auth: Kasper Jiang
# date: 2024-09-27 11:00:09
import logging
import os
from enum import Enum

import numpy as np
import pandas
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from . import file_util


class TypeKey(Enum):
    TEXT = 0
    HYPER_LINK = 1


class LinkKey(Enum):
    SHOW_TEXT = 'show_text'
    LINK = 'link'


class CellType:
    """
    支持文本、超链接
    注意：超链接的读取还不支持，所以不推荐使用超链接（虽然写入已支持）
    """

    def __init__(self, cell_type: TypeKey, cell_content):
        """
        存储类型和内容的类
        :param cell_type:
        :param cell_content: 当cell_type为超链接时，cell_content为字典，例如{LinkKey.TEXT:"点击地址", LinkKey.TEXT:"http://xxx", }
        """
        self.type = cell_type
        self.content = cell_content

    def __str__(self):
        return f"need to convet to {self.type}"


def get_col_names(sheet):
    """
    由于sheet.column_dimensions不支持遍历（但可通过[]访问），也不支持in操作符，所以将列名转为列表
    :param sheet:
    :return:
    """
    col_name_list = []
    if sheet is None:
        return col_name_list
    for column_cells in sheet.columns:
        col_letter = get_column_letter(column_cells[0].column)
        if col_letter:
            col_name_list.append(col_letter)
    return col_name_list


def read_excel_file(file_path, has_head=True):
    """
    读取excel文件并返回列表，元素为map，表头为map的key
    :param file_path:
    :param has_head: True=第一行作为key；False=列名作为key
    :return:元素为map的list
    """
    if not os.path.exists(file_path):
        return None

    if not has_head:
        header = None
    else:
        header = 0
    df = pandas.read_excel(file_path, header=header, dtype=object)
    # 将 NaN 替换为 None
    df = df.replace({np.nan: None})
    read_list = []

    for index, row in df.iterrows():
        data_dict = dict(zip(df.columns, row.tolist()))
        read_list.append(data_dict)
    return read_list


def write_excel_file(file_path, dict_list, keep_cell_format=True):
    """
    将数据保存为excel格式，支持以下功能：
    1. 单元格保存成超了解
    :param file_path: 含后缀的文件名，例如：“example.xlsx”
    :param dict_list: 元素为字典的列表，字典的key为列名，字典的值为单元格内容，单元格内容支持字符串、数字和自定义对象CellType（例如超链接），例如：
            [
            {'列1': '百度地址', '列2': CellType(TypeKey.HYPER_LINK,
                                             {LinkKey.SHOW_TEXT: '跳转百度', LinkKey.LINK: 'https://baidu.com'})},
            {'列1': '必应地址', '列2': CellType(TypeKey.HYPER_LINK,
                                             {LinkKey.SHOW_TEXT: '跳转必应', LinkKey.LINK: 'https://bing.com'})},
            ]
    :param keep_cell_format: 如果原文件存在，是否保留原表格单元格格式；只支持保留列宽
    :return:
    """
    if not file_path or not isinstance(dict_list, list):
        logging.error(f' file_path is None: {file_path} or \n dict_list is not a list: {dict_list}')
        return False
    file_util.create_parent_directory(file_path)

    # 读取原始 Excel 文件
    org_wb = None
    org_ws = None
    org_col_name_list = None
    if os.path.exists(file_path):
        org_wb = load_workbook(file_path)
        org_ws = org_wb.active
        org_col_name_list = get_col_names(org_ws)

    # 将数据保存到新的 Excel 文件
    df = pandas.DataFrame(dict_list)
    df.to_excel(file_path, index=False)
    dst_wb = load_workbook(file_path)
    dst_ws = dst_wb.active
    save_flag = False

    # 单元格类型处理
    for i, d in enumerate(dict_list):
        for j, (k, v) in enumerate(d.items()):
            if not isinstance(v, CellType):
                continue
            # 超链接进行特殊处理，其他类型都当作字符串或数字等默认类型
            if v.type == TypeKey.HYPER_LINK:
                # 行row和column都是从1开始索引的，第一行为列的标题，所以行从第2行开始
                cell = dst_ws.cell(row=i + 2, column=j + 1)
                cell.value = v.content[LinkKey.SHOW_TEXT]
                cell.hyperlink = v.content[LinkKey.LINK]
                cell.font = Font(color="0000FF", underline="single")  # 设置字体颜色为蓝色并添加下划线
                save_flag = True

    # 按原表格样式设置新表格
    if org_wb and org_ws:
        # 读取新 Excel 文件
        dst_col_name_list = get_col_names(dst_ws)
        # 复制原始列宽到新文件
        for column_cells in dst_ws.columns:
            column_letter = get_column_letter(column_cells[0].column)
            # 由于sheet.column_dimensions不支持遍历（但可通过[]访问），也不支持in操作符，所以将列名转为列表
            if column_letter in org_col_name_list and column_letter in dst_col_name_list:
                dst_ws.column_dimensions[column_letter].width = org_ws.column_dimensions[column_letter].width
                save_flag = True

    if save_flag:
        dst_wb.save(file_path)
    return True


def test_write_excel_file():
    file_path = '../../test/test_resource/test_write_excel_file.xlsx'
    data = [{'列1': '百度地址', '列2': CellType(TypeKey.HYPER_LINK,
                                                {LinkKey.SHOW_TEXT: '跳转百度', LinkKey.LINK: 'https://baidu.com'})},
            {'列1': '必应地址', '列2': CellType(TypeKey.HYPER_LINK,
                                                {LinkKey.SHOW_TEXT: '跳转必应', LinkKey.LINK: 'https://bing.com'})},
            ]
    write_excel_file(file_path, data)


if __name__ == '__main__':
    test_write_excel_file()
