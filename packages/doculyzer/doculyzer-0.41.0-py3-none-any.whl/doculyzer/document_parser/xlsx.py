"""
XLSX document parser module for the document pointer system.

This module parses Excel (XLSX) files into structured elements with comprehensive date extraction.
"""

import json
import logging
import os
import uuid
from typing import Dict, Any, List, Optional, Union, Tuple

from ..relationships import RelationshipType

try:
    # noinspection PyUnresolvedReferences
    import openpyxl

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logging.warning("openpyxl not available. Install with 'pip install openpyxl' to use XLSX parser")

from .base import DocumentParser
from .extract_dates import DateExtractor

logger = logging.getLogger(__name__)


class XlsxParser(DocumentParser):
    """Parser for Excel (XLSX) documents with enhanced date extraction."""

    def __init__(self, config: Optional[Dict[str, Any]] = None):
        """Initialize the XLSX parser."""
        super().__init__(config)

        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for XLSX parsing")

        # Configuration options
        self.config = config or {}
        self.extract_hidden_sheets = self.config.get("extract_hidden_sheets", False)
        self.extract_formulas = self.config.get("extract_formulas", True)
        self.extract_comments = self.config.get("extract_comments", True)
        self.extract_charts = self.config.get("extract_charts", False)
        self.extract_images = self.config.get("extract_images", False)
        self.max_rows = self.config.get("max_rows", 1000)  # Limit for large spreadsheets
        self.max_cols = self.config.get("max_cols", 100)  # Limit for very wide sheets
        self.temp_dir = self.config.get("temp_dir", os.path.join(os.path.dirname(__file__), 'temp'))
        self.max_content_preview = self.config.get("max_content_preview", 100)

        # Data table detection options
        self.detect_tables = self.config.get("detect_tables", True)  # Whether to detect data tables
        self.min_table_rows = self.config.get("min_table_rows", 2)  # Minimum rows for table detection
        self.min_table_cols = self.config.get("min_table_cols", 2)  # Minimum columns for table detection

        # Date extraction configuration
        self.extract_dates = self.config.get("extract_dates", True)
        self.date_context_chars = self.config.get("date_context_chars", 50)  # Small context window
        self.min_year = self.config.get("min_year", 1900)
        self.max_year = self.config.get("max_year", 2100)
        self.fiscal_year_start_month = self.config.get("fiscal_year_start_month", 10)
        self.default_locale = self.config.get("default_locale", "US")

        # Initialize date extractor if enabled
        self.date_extractor = None
        if self.extract_dates:
            try:
                self.date_extractor = DateExtractor(
                    context_chars=self.date_context_chars,
                    min_year=self.min_year,
                    max_year=self.max_year,
                    fiscal_year_start_month=self.fiscal_year_start_month,
                    default_locale=self.default_locale
                )
                logger.debug("Date extraction enabled with comprehensive temporal analysis for XLSX")
            except ImportError as e:
                logger.warning(f"Date extraction disabled: {e}")
                self.extract_dates = False

    """
    Fix the RGB object serialization issue in the Excel parser.
    """

    # Update the _extract_cell_style method to properly handle RGB objects
    @staticmethod
    def _extract_cell_style(cell) -> Dict[str, Any]:
        """
        Extract style information from an Excel cell with proper serialization.

        Args:
            cell: The openpyxl cell object

        Returns:
            Dictionary with style information
        """
        style = {}

        try:
            # Check if we're in read-only mode
            is_read_only = not hasattr(cell, 'font') or cell.font is None

            if is_read_only:
                # In read-only mode, we can't access style information
                return {}

            # Extract font properties
            if hasattr(cell, 'font') and cell.font:
                font = {}
                if hasattr(cell.font, 'bold') and cell.font.bold:
                    font["bold"] = True
                if hasattr(cell.font, 'italic') and cell.font.italic:
                    font["italic"] = True
                if hasattr(cell.font, 'underline') and cell.font.underline:
                    font["underline"] = True
                if hasattr(cell.font, 'strike') and cell.font.strike:
                    font["strike"] = True

                # Properly serialize RGB color
                if hasattr(cell.font, 'color') and cell.font.color:
                    if hasattr(cell.font.color, 'rgb') and cell.font.color.rgb:
                        # Convert RGB object to string if necessary
                        rgb_value = cell.font.color.rgb
                        if not isinstance(rgb_value, (str, type(None))):
                            # Convert to string to ensure it's serializable
                            rgb_value = str(rgb_value)
                        font["color"] = rgb_value

                if hasattr(cell.font, 'name') and cell.font.name:
                    font["name"] = cell.font.name
                if hasattr(cell.font, 'size') and cell.font.size:
                    font["size"] = cell.font.size

                if font:
                    style["font"] = font

            # Extract other style properties only if not in read-only mode
            if not is_read_only:
                # Extract alignment properties
                if hasattr(cell, 'alignment') and cell.alignment:
                    alignment = {}
                    props = [
                        'horizontal', 'vertical', 'textRotation', 'wrapText', 'shrinkToFit',
                        'indent', 'relativeIndent', 'justifyLastLine', 'readingOrder'
                    ]

                    for prop in props:
                        if hasattr(cell.alignment, prop):
                            value = getattr(cell.alignment, prop)
                            if value is not None:
                                # Convert camelCase to snake_case for property names
                                snake_prop = ''.join(['_' + c.lower() if c.isupper() else c for c in prop]).lstrip('_')
                                alignment[snake_prop] = value

                    if alignment:
                        style["alignment"] = alignment

                # Extract fill properties
                if hasattr(cell, 'fill') and cell.fill:
                    fill = {}
                    if hasattr(cell.fill, 'fill_type') and cell.fill.fill_type:
                        fill["type"] = cell.fill.fill_type

                    # Extract start (foreground) color with proper serialization
                    if hasattr(cell.fill, 'start_color') and cell.fill.start_color:
                        if hasattr(cell.fill.start_color, 'rgb') and cell.fill.start_color.rgb:
                            # Convert RGB object to string if necessary
                            rgb_value = cell.fill.start_color.rgb
                            if not isinstance(rgb_value, (str, type(None))):
                                # Convert to string to ensure it's serializable
                                rgb_value = str(rgb_value)
                            fill["start_color"] = rgb_value

                    # Extract end (background) color with proper serialization
                    if hasattr(cell.fill, 'end_color') and cell.fill.end_color:
                        if hasattr(cell.fill.end_color, 'rgb') and cell.fill.end_color.rgb:
                            # Convert RGB object to string if necessary
                            rgb_value = cell.fill.end_color.rgb
                            if not isinstance(rgb_value, (str, type(None))):
                                # Convert to string to ensure it's serializable
                                rgb_value = str(rgb_value)
                            fill["end_color"] = rgb_value

                    if fill:
                        style["fill"] = fill

                # Extract border properties
                if hasattr(cell, 'border') and cell.border:
                    border = {}

                    # Process each side of the border
                    for side in ['left', 'right', 'top', 'bottom', 'diagonal']:
                        side_border = getattr(cell.border, side, None)

                        if not side_border:
                            continue

                        side_info = {}

                        # Extract border style
                        if hasattr(side_border, 'style') and side_border.style:
                            side_info["style"] = side_border.style

                        # Extract border color with proper serialization
                        if hasattr(side_border, 'color') and side_border.color:
                            if hasattr(side_border.color, 'rgb') and side_border.color.rgb:
                                # Convert RGB object to string if necessary
                                rgb_value = side_border.color.rgb
                                if not isinstance(rgb_value, (str, type(None))):
                                    # Convert to string to ensure it's serializable
                                    rgb_value = str(rgb_value)
                                side_info["color"] = rgb_value

                        # Add to border dict if we have style info
                        if side_info:
                            if 'sides' not in border:
                                border["sides"] = {}
                            border["sides"][side] = side_info

                    # Add outline property if available
                    if hasattr(cell.border, 'outline') and cell.border.outline is not None:
                        border["outline"] = cell.border.outline

                    if border:
                        style["border"] = border

                # Extract number format
                if hasattr(cell, 'number_format') and cell.number_format:
                    style["number_format"] = cell.number_format

        except Exception as e:
            # Log error but continue processing
            logger.debug(f"Error extracting cell style: {str(e)}")

        return style

    # Add a helper method to ensure all values in metadata are JSON serializable
    def _ensure_serializable(self, obj):
        """
        Recursively check and convert non-serializable objects to serializable types.

        Args:
            obj: The object to check

        Returns:
            A JSON serializable version of the object
        """
        if isinstance(obj, dict):
            return {k: self._ensure_serializable(v) for k, v in obj.items()}
        elif isinstance(obj, list):
            return [self._ensure_serializable(item) for item in obj]
        elif isinstance(obj, tuple):
            return [self._ensure_serializable(item) for item in obj]
        elif isinstance(obj, (int, float, bool, str, type(None))):
            return obj
        else:
            # Convert any other type to string
            return str(obj)

    # Update the parse method to ensure serializable metadata and add date extraction
    def parse(self, doc_content: Dict[str, Any]) -> Dict[str, Any]:
        """
        Parse an XLSX document into structured elements with comprehensive date extraction.

        Args:
            doc_content: Document content and metadata

        Returns:
            Dictionary with document metadata, elements, relationships, extracted links, and dates
        """
        # Extract metadata from doc_content
        source_id = doc_content["id"]
        metadata = doc_content.get("metadata", {}).copy()  # Make a copy to avoid modifying original

        # Check if we have binary content or a path to a file
        binary_path = doc_content.get("binary_path")
        if not binary_path:
            # If we have binary content but no path, we need to save it to a temp file
            if not os.path.exists(self.temp_dir):
                os.makedirs(self.temp_dir, exist_ok=True)

            binary_content = doc_content.get("content", b"")
            if isinstance(binary_content, str):
                logger.warning("Expected binary content for XLSX but got string. Attempting to process anyway.")

            temp_file_path = os.path.join(self.temp_dir, f"temp_{uuid.uuid4().hex}.xlsx")
            with open(temp_file_path, 'wb') as f:
                if isinstance(binary_content, str):
                    f.write(binary_content.encode('utf-8'))
                else:
                    f.write(binary_content)

            binary_path = temp_file_path
            logger.debug(f"Saved binary content to temporary file: {binary_path}")

        # Generate document ID if not present
        doc_id = metadata.get("doc_id", self._generate_id("doc_"))

        # Load XLSX document
        try:
            # Use read_only mode for better performance with large files
            workbook = openpyxl.load_workbook(binary_path, read_only=True, data_only=not self.extract_formulas)
        except Exception as e:
            logger.error(f"Error loading XLSX document: {str(e)}")
            raise

        # Create document record with metadata
        doc_metadata = self._extract_document_metadata(workbook, metadata)

        # Ensure metadata is serializable
        doc_metadata = self._ensure_serializable(doc_metadata)

        document = {
            "doc_id": doc_id,
            "doc_type": "xlsx",
            "source": source_id,
            "metadata": doc_metadata,
            "content_hash": doc_content.get("content_hash", "")
        }

        # Create root element
        elements = [self._create_root_element(doc_id, source_id)]
        root_id = elements[0]["element_id"]

        # Initialize relationships list
        relationships = []

        # Parse document elements and create relationships
        workbook_elements, workbook_relationships = self._parse_workbook(workbook, doc_id, root_id, source_id)

        # Ensure all element metadata is serializable
        for element in workbook_elements:
            if "metadata" in element:
                element["metadata"] = self._ensure_serializable(element["metadata"])

        elements.extend(workbook_elements)
        relationships.extend(workbook_relationships)

        # Extract links from the document using the helper method
        links = self._extract_workbook_links(workbook, elements)

        # Extract dates from XLSX content with comprehensive temporal analysis
        element_dates = {}
        if self.extract_dates and self.date_extractor:
            try:
                # Extract dates from the entire workbook content
                full_text = self._extract_workbook_text(workbook)
                if full_text.strip():
                    document_dates = self.date_extractor.extract_dates_as_dicts(full_text)
                    if document_dates:
                        element_dates[root_id] = document_dates
                        logger.debug(f"Extracted {len(document_dates)} dates from XLSX document")

                # Extract dates from individual elements
                self._extract_dates_from_elements(elements, element_dates)

            except Exception as e:
                logger.warning(f"Error during XLSX date extraction: {e}")

        # Add date statistics to document metadata
        if element_dates:
            total_dates = sum(len(dates) for dates in element_dates.values())
            document["metadata"]["date_extraction"] = {
                "total_dates_found": total_dates,
                "elements_with_dates": len(element_dates),
                "extraction_enabled": True
            }
        else:
            document["metadata"]["date_extraction"] = {
                "total_dates_found": 0,
                "elements_with_dates": 0,
                "extraction_enabled": self.extract_dates
            }

        # Clean up temporary file if needed
        if binary_path != doc_content.get("binary_path") and os.path.exists(binary_path):
            try:
                os.remove(binary_path)
                logger.debug(f"Deleted temporary file: {binary_path}")
            except Exception as e:
                logger.warning(f"Failed to delete temporary file {binary_path}: {str(e)}")

        # Close workbook
        workbook.close()

        # Return the parsed document with extracted links, relationships, and dates
        result = {
            "document": document,
            "elements": elements,
            "links": links,
            "relationships": relationships
        }

        # Add dates if any were extracted
        if element_dates:
            result["element_dates"] = element_dates

        return result

    def _extract_workbook_text(self, workbook: openpyxl.workbook.Workbook) -> str:
        """
        Extract all text content from the workbook for document-level date extraction.

        Args:
            workbook: The Excel workbook

        Returns:
            Combined text from all sheets
        """
        all_text = []

        try:
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]

                # Skip hidden sheets if not configured to extract them
                if hasattr(sheet, 'sheet_state') and sheet.sheet_state == 'hidden' and not self.extract_hidden_sheets:
                    continue

                # Extract text from cells
                max_row = min(sheet.max_row or 0, self.max_rows)
                max_col = min(sheet.max_column or 0, self.max_cols)

                sheet_text = []
                for row_idx in range(1, max_row + 1):
                    for col_idx in range(1, max_col + 1):
                        cell = sheet.cell(row=row_idx, column=col_idx)
                        if cell.value is not None:
                            # Convert cell value to string for date extraction
                            cell_text = str(cell.value)
                            if cell_text.strip():
                                sheet_text.append(cell_text)

                # Add comments if available
                if self.extract_comments and hasattr(sheet, 'comments') and sheet.comments:
                    for comment in sheet.comments.values():
                        comment_text = comment.text if hasattr(comment, 'text') else str(comment)
                        if comment_text.strip():
                            sheet_text.append(comment_text)

                if sheet_text:
                    all_text.extend(sheet_text)

        except Exception as e:
            logger.debug(f"Error extracting workbook text for date analysis: {e}")

        return " ".join(all_text)

    def _extract_dates_from_elements(self, elements: List[Dict[str, Any]], element_dates: Dict[str, List[Dict[str, Any]]]):
        """
        Extract dates from individual XLSX elements.

        Args:
            elements: List of document elements
            element_dates: Dictionary to store extracted dates by element ID
        """
        if not self.date_extractor:
            return

        for element in elements:
            element_id = element.get("element_id")
            element_type = element.get("element_type", "")

            # Only extract dates from text-containing elements
            if element_type in ["sheet", "table_cell", "table_header", "comment",
                              "data_table", "table_header_row", "table_row_headers"]:

                try:
                    # Get the text content of this element
                    content_location = element.get("content_location", "{}")
                    location_data = json.loads(content_location)

                    # Extract text from the element
                    text_content = self._resolve_element_text(location_data)

                    if text_content and text_content.strip():
                        # Extract dates from this element's text
                        element_date_list = self.date_extractor.extract_dates_as_dicts(text_content)

                        if element_date_list:
                            element_dates[element_id] = element_date_list
                            logger.debug(f"Extracted {len(element_date_list)} dates from {element_type} element")

                except Exception as e:
                    logger.debug(f"Error extracting dates from element {element_id}: {e}")
                    continue

    """
    Fix all ReadOnly worksheet compatibility issues in the Excel parser.
    """

    # Fix the _process_rows method to handle ReadOnly worksheets
    def _process_rows(self, sheet, doc_id: str, sheet_id: str, source_id: str, max_row: int, max_col: int) -> Tuple[
        List[Dict[str, Any]], List[Dict[str, Any]]]:
        """
        Process rows from a worksheet.

        Args:
            sheet: The worksheet
            doc_id: Document ID
            sheet_id: Sheet element ID
            source_id: Source identifier
            max_row: Maximum row number
            max_col: Maximum column number

        Returns:
            Tuple of (list of elements, list of relationships)
        """
        elements = []
        relationships = []

        # Check if we're in read-only mode
        is_read_only = not hasattr(sheet, 'row_dimensions')

        # Define column letters for reference
        col_letters = [openpyxl.utils.get_column_letter(col_idx) for col_idx in range(1, max_col + 1)]

        # Process each row
        for row_idx in range(1, max_row + 1):
            # Check if row should be skipped
            # Checking if there is any cell in the row that has a value
            row_has_value = False
            for col_idx in range(1, max_col + 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                if cell.value is not None:
                    row_has_value = True
                    break

            if not row_has_value and self.config.get("skip_empty_rows", True):
                continue

            # Create row element
            row_id = self._generate_id(f"row_{row_idx}_")

            # Prepare row metadata with ReadOnly compatibility
            row_metadata = {
                "row": row_idx,
                "sheet": sheet.title
            }

            # Add height and hidden properties only if not in read-only mode
            if not is_read_only and hasattr(sheet, 'row_dimensions'):
                row_metadata["height"] = sheet.row_dimensions[
                    row_idx].height if row_idx in sheet.row_dimensions else None
                row_metadata["hidden"] = sheet.row_dimensions[
                    row_idx].hidden if row_idx in sheet.row_dimensions else False

            # Create row element
            row_element = {
                "element_id": row_id,
                "doc_id": doc_id,
                "element_type": "table_row",
                "parent_id": sheet_id,
                "content_preview": f"Row {row_idx}",
                "content_location": json.dumps({
                    "source": source_id,
                    "type": "table_row",
                    "sheet_name": sheet.title,
                    "row": row_idx
                }),
                "content_hash": "",
                "metadata": row_metadata
            }

            elements.append(row_element)

            # Create relationship from sheet to row
            contains_row_relationship = {
                "relationship_id": self._generate_id("rel_"),
                "source_id": sheet_id,
                "target_id": row_id,
                "relationship_type": RelationshipType.CONTAINS_TABLE_ROW.value,
                "metadata": {
                    "confidence": 1.0,
                    "row_index": row_idx
                }
            }
            relationships.append(contains_row_relationship)

            # Create inverse relationship
            row_contained_relationship = {
                "relationship_id": self._generate_id("rel_"),
                "source_id": row_id,
                "target_id": sheet_id,
                "relationship_type": RelationshipType.CONTAINED_BY.value,
                "metadata": {
                    "confidence": 1.0
                }
            }
            relationships.append(row_contained_relationship)

            # Process cells in this row
            for col_idx in range(1, max_col + 1):
                cell = sheet.cell(row=row_idx, column=col_idx)

                # Skip empty cells if requested
                if cell.value is None and self.config.get("skip_empty_cells", True):
                    continue

                # Get cell address (e.g., A1, B2)
                cell_addr = f"{col_letters[col_idx - 1]}{row_idx}"

                # Create cell element
                cell_id = self._generate_id(f"cell_{cell_addr}_")

                # Format cell value for display
                cell_value = cell.value
                if cell_value is not None:
                    if isinstance(cell_value, (int, float)):
                        content_preview = str(cell_value)
                    else:
                        content_preview = str(cell_value)
                else:
                    content_preview = ""

                # Limit preview length
                if len(content_preview) > self.max_content_preview:
                    content_preview = content_preview[:self.max_content_preview - 3] + "..."

                # Get cell style information (safely for both regular and ReadOnly worksheets)
                cell_style = self._extract_cell_style(cell)

                # Cell metadata
                cell_metadata = {
                    "address": cell_addr,
                    "row": row_idx,
                    "column": col_idx,
                    "column_letter": col_letters[col_idx - 1],
                    "data_type": cell.data_type if hasattr(cell, 'data_type') else None,
                    "style": cell_style
                }

                # Add formula if available and extracting formulas is enabled
                if self.extract_formulas and hasattr(cell, 'formula') and cell.formula:
                    cell_metadata["formula"] = cell.formula

                # Determine element type based on row position
                element_type = "table_header" if row_idx == 1 else "table_cell"

                # Create cell element
                cell_element = {
                    "element_id": cell_id,
                    "doc_id": doc_id,
                    "element_type": element_type,
                    "parent_id": row_id,
                    "content_preview": content_preview,
                    "content_location": json.dumps({
                        "source": source_id,
                        "type": element_type,
                        "sheet_name": sheet.title,
                        "cell": cell_addr
                    }),
                    "content_hash": self._generate_hash(content_preview),
                    "metadata": cell_metadata
                }

                elements.append(cell_element)

                # Create relationship from row to cell
                relationship_type = RelationshipType.CONTAINS_TABLE_HEADER.value if element_type == "table_header" else RelationshipType.CONTAINS_TABLE_CELL.value
                contains_cell_relationship = {
                    "relationship_id": self._generate_id("rel_"),
                    "source_id": row_id,
                    "target_id": cell_id,
                    "relationship_type": relationship_type,
                    "metadata": {
                        "confidence": 1.0,
                        "col_index": col_idx
                    }
                }
                relationships.append(contains_cell_relationship)

                # Create inverse relationship
                cell_contained_relationship = {
                    "relationship_id": self._generate_id("rel_"),
                    "source_id": cell_id,
                    "target_id": row_id,
                    "relationship_type": RelationshipType.CONTAINED_BY.value,
                    "metadata": {
                        "confidence": 1.0
                    }
                }
                relationships.append(cell_contained_relationship)

        return elements, relationships

    # Fix the _extract_merged_cells method to handle ReadOnly worksheets
    def _extract_merged_cells(self, sheet, doc_id: str, sheet_id: str, source_id: str) -> Tuple[
        List[Dict[str, Any]], List[Dict[str, Any]]]:
        """
        Extract merged cells information.

        Args:
            sheet: The worksheet
            doc_id: Document ID
            sheet_id: Sheet element ID
            source_id: Source identifier

        Returns:
            Tuple of (list of merged cell elements, list of relationships)
        """
        elements = []
        relationships = []

        # Check if sheet has merged_cells attribute (only in regular worksheets, not ReadOnly)
        if not hasattr(sheet, 'merged_cells') or not sheet.merged_cells:
            return elements, relationships

        # Create merged cells container element
        merged_cells_id = self._generate_id("merged_cells_")
        merged_cells_element = {
            "element_id": merged_cells_id,
            "doc_id": doc_id,
            "element_type": "merged_cells",
            "parent_id": sheet_id,
            "content_preview": f"Merged cells in sheet '{sheet.title}'",
            "content_location": json.dumps({
                "source": source_id,
                "type": "merged_cells",
                "sheet_name": sheet.title
            }),
            "content_hash": "",
            "metadata": {
                "sheet": sheet.title,
                "count": len(sheet.merged_cells.ranges) if hasattr(sheet.merged_cells, 'ranges') else 0
            }
        }
        elements.append(merged_cells_element)

        # Create relationship from sheet to merged cells container
        contains_merged_cells_relationship = {
            "relationship_id": self._generate_id("rel_"),
            "source_id": sheet_id,
            "target_id": merged_cells_id,
            "relationship_type": RelationshipType.CONTAINS.value,
            "metadata": {
                "confidence": 1.0
            }
        }
        relationships.append(contains_merged_cells_relationship)

        # Create inverse relationship
        merged_cells_contained_relationship = {
            "relationship_id": self._generate_id("rel_"),
            "source_id": merged_cells_id,
            "target_id": sheet_id,
            "relationship_type": RelationshipType.CONTAINED_BY.value,
            "metadata": {
                "confidence": 1.0
            }
        }
        relationships.append(merged_cells_contained_relationship)

        # Process each merged cell range
        for merged_idx, merged_range in enumerate(sheet.merged_cells.ranges):
            # Create merged cell element
            merged_id = self._generate_id(f"merged_{merged_idx}_")

            # Get coordinate information
            min_row, min_col, max_row, max_col = merged_range.min_row, merged_range.min_col, merged_range.max_row, merged_range.max_col

            # Get the value from the top-left cell of the merge
            top_left_cell = sheet.cell(row=min_row, column=min_col)
            value = top_left_cell.value
            content_preview = str(value) if value is not None else ""

            # Limit preview length
            if len(content_preview) > self.max_content_preview:
                content_preview = content_preview[:self.max_content_preview - 3] + "..."

            # Create merged cell element
            merged_element = {
                "element_id": merged_id,
                "doc_id": doc_id,
                "element_type": "merged_cell",
                "parent_id": merged_cells_id,
                "content_preview": content_preview,
                "content_location": json.dumps({
                    "source": source_id,
                    "type": "merged_cell",
                    "sheet_name": sheet.title,
                    "range": str(merged_range)
                }),
                "content_hash": self._generate_hash(content_preview),
                "metadata": {
                    "range": str(merged_range),
                    "min_row": min_row,
                    "min_column": min_col,
                    "max_row": max_row,
                    "max_column": max_col,
                    "sheet": sheet.title
                }
            }

            elements.append(merged_element)

            # Create relationship from merged cells container to merged cell
            contains_merged_cell_relationship = {
                "relationship_id": self._generate_id("rel_"),
                "source_id": merged_cells_id,
                "target_id": merged_id,
                "relationship_type": RelationshipType.CONTAINS.value,
                "metadata": {
                    "confidence": 1.0,
                    "index": merged_idx
                }
            }
            relationships.append(contains_merged_cell_relationship)

            # Create inverse relationship
            merged_cell_contained_relationship = {
                "relationship_id": self._generate_id("rel_"),
                "source_id": merged_id,
                "target_id": merged_cells_id,
                "relationship_type": RelationshipType.CONTAINED_BY.value,
                "metadata": {
                    "confidence": 1.0
                }
            }
            relationships.append(merged_cell_contained_relationship)

        return elements, relationships

    # Fix the _extract_comments method to handle ReadOnly worksheets
    def _extract_comments(self, sheet, doc_id: str, sheet_id: str, source_id: str) -> Tuple[
        List[Dict[str, Any]], List[Dict[str, Any]]]:
        """
        Extract comments from worksheet.

        Args:
            sheet: The worksheet
            doc_id: Document ID
            sheet_id: Sheet element ID
            source_id: Source identifier

        Returns:
            Tuple of (list of comment elements, list of relationships)
        """
        elements = []
        relationships = []

        # Check if there are comments (comments are not accessible in ReadOnly mode)
        if not hasattr(sheet, 'comments') or not sheet.comments:
            return elements, relationships

        # Create comments container element
        comments_id = self._generate_id("comments_")
        comments_element = {
            "element_id": comments_id,
            "doc_id": doc_id,
            "element_type": "comments",
            "parent_id": sheet_id,
            "content_preview": f"Comments in sheet '{sheet.title}'",
            "content_location": json.dumps({
                "source": source_id,
                "type": "comments",
                "sheet_name": sheet.title
            }),
            "content_hash": "",
            "metadata": {
                "sheet": sheet.title,
                "count": len(sheet.comments)
            }
        }
        elements.append(comments_element)

        # Create relationship from sheet to comments container
        contains_comments_relationship = {
            "relationship_id": self._generate_id("rel_"),
            "source_id": sheet_id,
            "target_id": comments_id,
            "relationship_type": RelationshipType.CONTAINS.value,
            "metadata": {
                "confidence": 1.0
            }
        }
        relationships.append(contains_comments_relationship)

        # Create inverse relationship
        comments_contained_relationship = {
            "relationship_id": self._generate_id("rel_"),
            "source_id": comments_id,
            "target_id": sheet_id,
            "relationship_type": RelationshipType.CONTAINED_BY.value,
            "metadata": {
                "confidence": 1.0
            }
        }
        relationships.append(comments_contained_relationship)

        # Process comments
        for comment_idx, (cell_addr, comment) in enumerate(sheet.comments.items()):
            # Create comment element
            comment_id = self._generate_id(f"comment_{comment_idx}_")

            # Extract comment text and author
            text = comment.text if hasattr(comment, 'text') else str(comment)
            author = comment.author if hasattr(comment, 'author') else "Unknown"

            # Limit text length for preview
            content_preview = f"Comment by {author}: {text}"
            if len(content_preview) > self.max_content_preview:
                content_preview = content_preview[:self.max_content_preview - 3] + "..."

            # Create comment element
            comment_element = {
                "element_id": comment_id,
                "doc_id": doc_id,
                "element_type": "comment",
                "parent_id": comments_id,
                "content_preview": content_preview,
                "content_location": json.dumps({
                    "source": source_id,
                    "type": "comment",
                    "sheet_name": sheet.title,
                    "cell": cell_addr
                }),
                "content_hash": self._generate_hash(text),
                "metadata": {
                    "cell": cell_addr,
                    "author": author,
                    "text": text,
                    "sheet": sheet.title
                }
            }

            elements.append(comment_element)

            # Create relationship from comments container to comment
            contains_comment_relationship = {
                "relationship_id": self._generate_id("rel_"),
                "source_id": comments_id,
                "target_id": comment_id,
                "relationship_type": RelationshipType.CONTAINS.value,
                "metadata": {
                    "confidence": 1.0,
                    "cell": cell_addr
                }
            }
            relationships.append(contains_comment_relationship)

            # Create inverse relationship
            comment_contained_relationship = {
                "relationship_id": self._generate_id("rel_"),
                "source_id": comment_id,
                "target_id": comments_id,
                "relationship_type": RelationshipType.CONTAINED_BY.value,
                "metadata": {
                    "confidence": 1.0
                }
            }
            relationships.append(comment_contained_relationship)

        return elements, relationships

    # Update the _parse_workbook method to handle ReadOnly worksheets
    def _parse_workbook(self, workbook: openpyxl.workbook.Workbook, doc_id: str, parent_id: str, source_id: str) -> \
            Tuple[
                List[Dict[str, Any]], List[Dict[str, Any]]]:
        """
        Parse Excel workbook into structured elements.

        Args:
            workbook: The openpyxl Workbook
            doc_id: Document ID
            parent_id: Parent element ID
            source_id: Source identifier

        Returns:
            Tuple of (list of elements, list of relationships)
        """
        elements = []
        relationships = []

        # Create workbook element
        workbook_id = self._generate_id("workbook_")
        workbook_element = {
            "element_id": workbook_id,
            "doc_id": doc_id,
            "element_type": "workbook",
            "parent_id": parent_id,
            "content_preview": f"Excel workbook with {len(workbook.sheetnames)} sheets",
            "content_location": json.dumps({
                "source": source_id,
                "type": "workbook"
            }),
            "content_hash": "",
            "metadata": {
                "sheet_count": len(workbook.sheetnames),
                "active_sheet": workbook.active.title if hasattr(workbook, 'active') and workbook.active else None
            }
        }

        elements.append(workbook_element)

        # Create relationship from root to workbook
        contains_relationship = {
            "relationship_id": self._generate_id("rel_"),
            "source_id": parent_id,
            "target_id": workbook_id,
            "relationship_type": RelationshipType.CONTAINS.value,
            "metadata": {
                "confidence": 1.0
            }
        }
        relationships.append(contains_relationship)

        # Create inverse relationship
        contained_by_relationship = {
            "relationship_id": self._generate_id("rel_"),
            "source_id": workbook_id,
            "target_id": parent_id,
            "relationship_type": RelationshipType.CONTAINED_BY.value,
            "metadata": {
                "confidence": 1.0
            }
        }
        relationships.append(contained_by_relationship)

        # Process each sheet
        for sheet_idx, sheet_name in enumerate(workbook.sheetnames):
            sheet = workbook[sheet_name]

            # Skip hidden sheets if not configured to extract them
            # ReadOnly worksheets may not have sheet_state attribute
            if hasattr(sheet, 'sheet_state') and sheet.sheet_state == 'hidden' and not self.extract_hidden_sheets:
                logger.debug(f"Skipping hidden sheet: {sheet_name}")
                continue

            # Process the sheet
            sheet_elements, sheet_relationships = self._process_sheet(sheet, doc_id, workbook_id, source_id, sheet_idx)
            elements.extend(sheet_elements)
            relationships.extend(sheet_relationships)

        return elements, relationships

    # Add the _process_sheet method with sheet_idx parameter
    def _process_sheet(self, sheet, doc_id: str, parent_id: str, source_id: str, sheet_idx: int = 0) -> Tuple[
        List[Dict[str, Any]], List[Dict[str, Any]]]:
        """
        Process a single worksheet.

        Args:
            sheet: The worksheet
            doc_id: Document ID
            parent_id: Parent element ID
            source_id: Source identifier
            sheet_idx: Index of sheet in workbook (0-based)

        Returns:
            Tuple of (list of sheet elements, list of relationships)
        """
        elements = []
        relationships = []

        # Create sheet element
        sheet_id = self._generate_id("sheet_")

        # Get sheet dimensions
        max_row = min(sheet.max_row or 0, self.max_rows)
        max_col = min(sheet.max_column or 0, self.max_cols)

        # Create sheet preview
        if max_row > 0 and max_col > 0:
            preview = f"Sheet '{sheet.title}' with {max_row} rows and {max_col} columns"
        else:
            preview = f"Empty sheet '{sheet.title}'"

        # Sheet metadata
        sheet_metadata = {
            "title": sheet.title,
            "max_row": max_row,
            "max_column": max_col
        }

        # Add sheet state if available
        if hasattr(sheet, 'sheet_state'):
            sheet_metadata["sheet_state"] = sheet.sheet_state
        else:
            # Default to visible for read-only worksheets
            sheet_metadata["sheet_state"] = "visible"

        # Check if sheet has autofilter (safely for read-only worksheets)
        if hasattr(sheet, 'auto_filter') and sheet.auto_filter:
            sheet_metadata["has_autofilter"] = True
            sheet_metadata["autofilter_range"] = str(sheet.auto_filter.ref) if hasattr(sheet.auto_filter,
                                                                                       'ref') else None

        # Check if sheet has freeze panes (safely for read-only worksheets)
        if hasattr(sheet, 'freeze_panes') and sheet.freeze_panes:
            sheet_metadata["has_freeze_panes"] = True
            sheet_metadata["freeze_panes"] = str(sheet.freeze_panes)

        # Create sheet element
        sheet_element = {
            "element_id": sheet_id,
            "doc_id": doc_id,
            "element_type": "sheet",
            "parent_id": parent_id,
            "content_preview": preview,
            "content_location": json.dumps({
                "source": source_id,
                "type": "sheet",
                "sheet_name": sheet.title
            }),
            "content_hash": self._generate_hash(preview),
            "metadata": sheet_metadata
        }

        elements.append(sheet_element)

        # Create relationship from workbook to sheet
        contains_sheet_relationship = {
            "relationship_id": self._generate_id("rel_"),
            "source_id": parent_id,
            "target_id": sheet_id,
            "relationship_type": RelationshipType.CONTAINS.value,
            "metadata": {
                "confidence": 1.0,
                "index": sheet_idx
            }
        }
        relationships.append(contains_sheet_relationship)

        # Create inverse relationship
        sheet_contained_relationship = {
            "relationship_id": self._generate_id("rel_"),
            "source_id": sheet_id,
            "target_id": parent_id,
            "relationship_type": RelationshipType.CONTAINED_BY.value,
            "metadata": {
                "confidence": 1.0
            }
        }
        relationships.append(sheet_contained_relationship)

        # Extract sheet structure
        if max_row > 0 and max_col > 0:
            # Process rows
            row_elements, row_relationships = self._process_rows(sheet, doc_id, sheet_id, source_id, max_row, max_col)
            elements.extend(row_elements)
            relationships.extend(row_relationships)

            # Detect and extract data tables if enabled
            if self.detect_tables and max_row >= self.min_table_rows and max_col >= self.min_table_cols:
                data_table_elements, data_table_relationships = self._detect_data_tables(sheet, doc_id, sheet_id,
                                                                                         source_id, max_row, max_col)
                elements.extend(data_table_elements)
                relationships.extend(data_table_relationships)

            # Extract merged cells (if available)
            if hasattr(sheet, 'merged_cells') and sheet.merged_cells:
                merged_elements, merged_relationships = self._extract_merged_cells(sheet, doc_id, sheet_id, source_id)
                elements.extend(merged_elements)
                relationships.extend(merged_relationships)

            # Extract comments (if available and enabled)
            if self.extract_comments and hasattr(sheet, 'comments') and sheet.comments:
                comment_elements, comment_relationships = self._extract_comments(sheet, doc_id, sheet_id, source_id)
                elements.extend(comment_elements)
                relationships.extend(comment_relationships)

        return elements, relationships

    # Add necessary helper methods
    def _create_root_element(self, doc_id, source_id):
        """
        Create the root element for the document.

        Args:
            doc_id: Document ID
            source_id: Source identifier

        Returns:
            Root element dictionary
        """
        return {
            "element_id": self._generate_id("root_"),
            "doc_id": doc_id,
            "element_type": "root",
            "parent_id": None,
            "content_preview": "Document Root",
            "content_location": json.dumps({
                "source": source_id,
                "type": "root"
            }),
            "content_hash": "",
            "metadata": {}
        }

    def _generate_id(self, prefix="id_"):
        """
        Generate a unique ID.

        Args:
            prefix: Prefix for the ID

        Returns:
            Generated ID string
        """
        import uuid
        return f"{prefix}{uuid.uuid4().hex}"

    def _generate_hash(self, content):
        """
        Generate a hash for content.

        Args:
            content: Content to hash

        Returns:
            Hash string
        """
        import hashlib

        if isinstance(content, str):
            content_bytes = content.encode('utf-8')
        elif isinstance(content, bytes):
            content_bytes = content
        else:
            content_bytes = str(content).encode('utf-8')

        return hashlib.sha256(content_bytes).hexdigest()

    """
    Add the missing methods and fix the index attribute issue in the Excel parser.
    """

    # Modify the _process_sheet function to fix the issue with sheet.index
    # In the _parse_workbook function, around line 554 in xlsx.py

    def _resolve_element_text(self, location_data: Dict[str, Any], source_content: Optional[Union[str, bytes]] = None) -> str:
        """
        Resolve the plain text representation of an Excel element.

        Args:
            location_data: Content location data
            source_content: Optional preloaded source content

        Returns:
            Plain text representation of the element
        """
        # Get the content using the existing method
        content = self._resolve_element_content(location_data, source_content)
        element_type = location_data.get("type", "")

        # Handle specific element types
        if element_type == "workbook":
            # For workbook, return a simple summary
            return content.strip()

        elif element_type == "sheet":
            # For sheets, return basic sheet information
            return content.strip()

        elif element_type == "table_row":
            # For rows, the content is already tab-separated values
            return content.strip()

        elif element_type == "table_cell" or element_type == "table_header":
            # For cells, just return the cell value
            return content.strip()

        elif element_type == "data_table":
            # For data tables, preserve the structure but remove any Excel-specific formatting
            # The _resolve_element_content method already returns structured text
            return content.strip()

        elif element_type == "comment":
            # For comments, extract just the comment text without the author info
            if "Comment by " in content:
                # Extract comment text after the author prefix
                return content.split(": ", 1)[1].strip()
            return content.strip()

        elif element_type == "merged_cell":
            # For merged cells, return just the cell value
            return content.strip()

        # Default: return the content as is
        return content.strip()

    def _resolve_element_content(self, location_data: Dict[str, Any],
                                 source_content: Optional[Union[str, bytes]]) -> str:
        """
        Resolve content for specific XLSX element types.

        Args:
            location_data: Content location data
            source_content: Optional pre-loaded source content

        Returns:
            Resolved content string
        """
        source = location_data.get("source", "")
        element_type = location_data.get("type", "")
        sheet_name = location_data.get("sheet_name", "")

        # Load the document if source content is not provided
        wb = None
        temp_file = None
        try:
            if source_content is None:
                # Check if source is a file path
                if os.path.exists(source):
                    try:
                        wb = openpyxl.load_workbook(source, read_only=True, data_only=not self.extract_formulas)
                    except Exception as e:
                        raise ValueError(f"Error loading XLSX document: {str(e)}")
                else:
                    raise ValueError(f"Source file not found: {source}")
            else:
                # Save content to a temporary file
                if not os.path.exists(self.temp_dir):
                    os.makedirs(self.temp_dir, exist_ok=True)

                import uuid
                temp_file = os.path.join(self.temp_dir, f"temp_{uuid.uuid4().hex}.xlsx")
                with open(temp_file, 'wb') as f:
                    if isinstance(source_content, str):
                        f.write(source_content.encode('utf-8'))
                    else:
                        f.write(source_content)

                # Load the document
                try:
                    wb = openpyxl.load_workbook(temp_file, read_only=True, data_only=not self.extract_formulas)
                except Exception as e:
                    raise ValueError(f"Error loading XLSX document: {str(e)}")

            # Handle different element types
            if element_type == "workbook":
                # Return information about the workbook
                sheet_names = wb.sheetnames
                active_sheet = wb.active.title if hasattr(wb, 'active') and wb.active else None
                return f"Workbook with sheets: {', '.join(sheet_names)}. Active sheet: {active_sheet or 'None'}"

            # Check if sheet exists
            if sheet_name and sheet_name not in wb.sheetnames:
                raise ValueError(f"Sheet '{sheet_name}' not found in workbook")

            # Get the specified sheet
            if sheet_name:
                sheet = wb[sheet_name]
            else:
                # Use active sheet if no specific sheet name provided
                sheet = wb.active

            if element_type == "sheet":
                # Return information about the sheet
                max_row = min(sheet.max_row or 0, self.max_rows)
                max_col = min(sheet.max_column or 0, self.max_cols)
                return f"Sheet '{sheet.title}' with {max_row} rows and {max_col} columns"

            elif element_type == "table_row":
                # Extract row by index
                row = location_data.get("row", 0)

                if row <= 0 or row > min(sheet.max_row or 0, self.max_rows):
                    return f"Row {row} is out of range"

                row_values = []
                for col in range(1, min(sheet.max_column + 1, self.max_cols + 1)):
                    cell = sheet.cell(row=row, column=col)
                    row_values.append(str(cell.value) if cell.value is not None else "")

                return "\t".join(row_values)

            elif element_type == "table_cell":
                # Extract cell by reference
                cell_ref = location_data.get("cell", "")

                if cell_ref:
                    # Direct cell reference (e.g., "A1")
                    try:
                        cell = sheet[cell_ref]
                        return str(cell.value) if cell.value is not None else ""
                    except Exception as e:
                        return f"Error accessing cell {cell_ref}: {str(e)}"
                else:
                    # Row/column coordinates
                    row = location_data.get("row", 0)
                    col = location_data.get("col", 0)

                    if row <= 0 or col <= 0 or row > min(sheet.max_row or 0, self.max_rows) or col > min(
                            sheet.max_column or 0, self.max_cols):
                        return f"Cell at row {row}, column {col} is out of range"

                    cell = sheet.cell(row=row, column=col)
                    return str(cell.value) if cell.value is not None else ""

            elif element_type == "data_table":
                # Extract a range of cells forming a table
                range_str = location_data.get("range", "")

                if not range_str:
                    return "No range specified for data table"

                try:
                    # Get all cells in the range
                    from openpyxl.utils.cell import range_boundaries
                    min_col, min_row, max_col, max_row = range_boundaries(range_str)

                    # Extract table data
                    table_data = []
                    for row in range(min_row, max_row + 1):
                        row_data = []
                        for col in range(min_col, max_col + 1):
                            cell = sheet.cell(row=row, column=col)
                            row_data.append(str(cell.value) if cell.value is not None else "")
                        table_data.append("\t".join(row_data))

                    return "\n".join(table_data)
                except Exception as e:
                    return f"Error extracting data table: {str(e)}"

            elif element_type == "comment":
                # Extract comment from cell
                cell_ref = location_data.get("cell", "")

                if not cell_ref or not hasattr(sheet, "comments") or not sheet.comments:
                    return "No comment found"

                if cell_ref in sheet.comments:
                    comment = sheet.comments[cell_ref]
                    author = comment.author if hasattr(comment, 'author') else "Unknown"
                    text = comment.text if hasattr(comment, 'text') else str(comment)
                    return f"Comment by {author}: {text}"
                else:
                    return f"No comment found at cell {cell_ref}"

            elif element_type == "merged_cell":
                # Extract merged cell content
                range_str = location_data.get("range", "")

                if not range_str:
                    return "No range specified for merged cell"

                try:
                    from openpyxl.utils.cell import range_boundaries
                    min_col, min_row, max_col, max_row = range_boundaries(range_str)

                    # Get top-left cell (contains the value for merged cells)
                    cell = sheet.cell(row=min_row, column=min_col)
                    return str(cell.value) if cell.value is not None else ""
                except Exception as e:
                    return f"Error extracting merged cell content: {str(e)}"

            else:
                # Default: return the sheet content as text
                max_row = min(sheet.max_row or 0, self.max_rows)
                max_col = min(sheet.max_column or 0, self.max_cols)

                sheet_content = []
                for row in range(1, max_row + 1):
                    row_values = []
                    for col in range(1, max_col + 1):
                        cell = sheet.cell(row=row, column=col)
                        row_values.append(str(cell.value) if cell.value is not None else "")
                    sheet_content.append("\t".join(row_values))

                return "\n".join(sheet_content)

        finally:
            # Close workbook
            if wb:
                wb.close()

            # Clean up temporary file
            if temp_file and os.path.exists(temp_file):
                try:
                    os.remove(temp_file)
                except Exception as e:
                    logger.warning(f"Failed to delete temporary file {temp_file}: {str(e)}")

    def supports_location(self, content_location: Dict[str, any]) -> bool:
        """
        Check if this parser supports resolving the given location.

        Args:
            content_location: Content location pointer

        Returns:
            True if supported, False otherwise
        """
        try:
            location_data = content_location
            source = location_data.get("source", "")

            # Check if source exists and is a file
            if not os.path.exists(source) or not os.path.isfile(source):
                return False

            # Check file extension for XLSX
            _, ext = os.path.splitext(source.lower())
            return ext in ['.xlsx', '.xlsm', '.xltx', '.xltm']

        except (json.JSONDecodeError, TypeError):
            return False

    @staticmethod
    def _extract_document_metadata(workbook: openpyxl.workbook.Workbook, base_metadata: Dict[str, Any]) -> Dict[
        str, Any]:
        """
        Extract metadata from XLSX document.

        Args:
            workbook: The XLSX workbook
            base_metadata: Base metadata from content source

        Returns:
            Dictionary of document metadata
        """
        # Combine base metadata with document properties
        metadata = base_metadata.copy()

        try:
            # Get workbook properties
            props = workbook.properties

            # Add core properties to metadata
            if props.title:
                metadata["title"] = props.title
            if props.creator:
                metadata["author"] = props.creator
            if props.created:
                metadata["created"] = props.created.timestamp() if hasattr(props.created, 'timestamp') else str(
                    props.created)
            if props.modified:
                metadata["modified"] = props.modified.timestamp() if hasattr(props.modified, 'timestamp') else str(
                    props.modified)
            if props.lastModifiedBy:
                metadata["last_modified_by"] = props.lastModifiedBy
            if props.subject:
                metadata["subject"] = props.subject
            if props.keywords:
                metadata["keywords"] = props.keywords
            if props.category:
                metadata["category"] = props.category
            if props.description:
                metadata["description"] = props.description

            # Add document statistics
            metadata["sheet_count"] = len(workbook.sheetnames)
            metadata["sheet_names"] = workbook.sheetnames

            # Add calculation properties if available
            calc_props = {}
            if hasattr(workbook, 'calculation') and workbook.calculation:
                calc = workbook.calculation
                if hasattr(calc, 'calcMode'):
                    calc_props["calc_mode"] = calc.calcMode
                if hasattr(calc, 'calcCompleted'):
                    calc_props["calc_completed"] = calc.calcCompleted
                if hasattr(calc, 'calcOnSave'):
                    calc_props["calc_on_save"] = calc.calcOnSave
                if hasattr(calc, 'fullCalcOnLoad'):
                    calc_props["full_calc_on_load"] = calc.fullCalcOnLoad

            if calc_props:
                metadata["calculation"] = calc_props

        except Exception as e:
            logger.warning(f"Error extracting document metadata: {str(e)}")

        return metadata

    def _detect_data_tables(self, sheet, doc_id: str, sheet_id: str, source_id: str, max_row: int, max_col: int) -> \
            Tuple[
                List[Dict[str, Any]], List[Dict[str, Any]]]:
        """
        Detect and extract structured data tables within a worksheet.
        This focuses on finding regions that appear to be 2D tables with headers.

        Args:
            sheet: The worksheet
            doc_id: Document ID
            sheet_id: Sheet element ID
            source_id: Source identifier
            max_row: Maximum row number
            max_col: Maximum column number

        Returns:
            Tuple of (list of data table elements, list of relationships)
        """
        elements = []
        relationships = []

        # Skip small sheets that are unlikely to contain meaningful tables
        if max_row < self.min_table_rows or max_col < self.min_table_cols:
            return elements, relationships

        # Use heuristics to detect potential tables
        # 1. Check for consistent data regions with header rows
        # 2. Look for formatting patterns (e.g., header row formatting)
        # 3. Check for the presence of autofilters (strong indicator of tables)

        # Start by detecting candidate table regions
        table_regions = []

        # If an autofilter is present, it's a strong indication of a table
        if hasattr(sheet, 'auto_filter') and sheet.auto_filter and hasattr(sheet.auto_filter, 'ref'):
            autofilter_range = sheet.auto_filter.ref
            if autofilter_range:
                # Parse the range (e.g., "A1:F20")
                try:
                    # Get the coordinates from the range
                    from openpyxl.utils.cell import range_boundaries
                    min_col, min_row, max_col, max_row = range_boundaries(autofilter_range)

                    table_regions.append({
                        "min_row": min_row,
                        "min_col": min_col,
                        "max_row": max_row,
                        "max_col": max_col,
                        "has_header": True,  # Assume first row is header in autofilter
                        "confidence": "high"  # High confidence due to autofilter
                    })
                except Exception as e:
                    logger.debug(f"Error parsing autofilter range: {str(e)}")

        # If no autofilter, try to detect tables by analyzing data patterns
        if not table_regions:
            # Get a snapshot of the data
            data_snapshot = []
            for row_idx in range(1, min(max_row + 1, 20)):  # Limit to first 20 rows for performance
                row_data = []
                empty_count = 0
                for col_idx in range(1, min(max_col + 1, 20)):  # Limit to first 20 columns for performance
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    is_empty = cell.value is None
                    row_data.append({
                        "value": cell.value,
                        "is_empty": is_empty,
                        "is_bold": hasattr(cell, 'font') and cell.font and hasattr(cell.font,
                                                                                   'bold') and cell.font.bold,
                        "has_fill": hasattr(cell, 'fill') and cell.fill and hasattr(cell.fill, 'fill_type') and
                                    cell.fill.fill_type != 'none'
                    })
                    if is_empty:
                        empty_count += 1

                # Skip completely empty rows
                if empty_count == len(row_data):
                    continue

                data_snapshot.append(row_data)

            # Simple table detection: Look for patterns like:
            # - First row has formatting different from other rows (likely headers)
            # - Consistent data in columns
            # - Few empty cells in the middle of data

            if len(data_snapshot) >= 2:  # Need at least 2 rows (header + data)
                first_row = data_snapshot[0]

                # Check if first row might be a header row
                header_indicators = 0
                for cell_data in first_row:
                    if cell_data["is_bold"] or cell_data["has_fill"]:
                        header_indicators += 1

                # If more than half of cells in first row have header-like formatting
                is_likely_header = header_indicators > len(first_row) / 2

                # Find how many columns might be in the table
                max_col_with_data = 0
                for row_data in data_snapshot:
                    for col_idx, cell_data in enumerate(row_data):
                        if not cell_data["is_empty"]:
                            max_col_with_data = max(max_col_with_data, col_idx + 1)

                # Check for consistent data patterns in columns
                if max_col_with_data >= 2:  # Need at least 2 columns
                    # Calculate approximate table boundaries
                    table_min_row = 1
                    table_min_col = 1

                    # Determine table height: look for empty rows or significant formatting changes
                    table_max_row = len(data_snapshot)
                    for row_idx in range(1, len(data_snapshot)):
                        empty_count = sum(1 for cell in data_snapshot[row_idx] if cell["is_empty"])
                        if empty_count > max_col_with_data / 2:  # Over half the cells are empty
                            # This might be the end of the table
                            table_max_row = row_idx
                            break

                    table_max_col = max_col_with_data

                    # If we found a viable table region, add it
                    if table_max_row >= 2 and table_max_col >= 2:
                        table_regions.append({
                            "min_row": table_min_row,
                            "min_col": table_min_col,
                            "max_row": table_max_row,
                            "max_col": table_max_col,
                            "has_header": is_likely_header,
                            "confidence": "medium"  # Medium confidence due to heuristics
                        })

        # Create data tables container
        if table_regions:
            data_tables_id = self._generate_id("data_tables_")
            data_tables_element = {
                "element_id": data_tables_id,
                "doc_id": doc_id,
                "element_type": "data_tables",
                "parent_id": sheet_id,
                "content_preview": f"Data tables in sheet '{sheet.title}'",
                "content_location": json.dumps({
                    "source": source_id,
                    "type": "data_tables",
                    "sheet_name": sheet.title
                }),
                "content_hash": "",
                "metadata": {
                    "sheet": sheet.title,
                    "count": len(table_regions)
                }
            }
            elements.append(data_tables_element)

            # Create relationship from sheet to data tables container
            contains_data_tables_relationship = {
                "relationship_id": self._generate_id("rel_"),
                "source_id": sheet_id,
                "target_id": data_tables_id,
                "relationship_type": RelationshipType.CONTAINS.value,
                "metadata": {
                    "confidence": 1.0
                }
            }
            relationships.append(contains_data_tables_relationship)

            # Create inverse relationship
            data_tables_contained_relationship = {
                "relationship_id": self._generate_id("rel_"),
                "source_id": data_tables_id,
                "target_id": sheet_id,
                "relationship_type": RelationshipType.CONTAINED_BY.value,
                "metadata": {
                    "confidence": 1.0
                }
            }
            relationships.append(data_tables_contained_relationship)

            # Process each detected table region
            for idx, region in enumerate(table_regions):
                # Create a data table element
                table_id = self._generate_id(f"data_table_{idx + 1}_")

                # Generate a table preview with header row if present
                preview = f"Data table {idx + 1} ({region['max_row'] - region['min_row'] + 1}x{region['max_col'] - region['min_col'] + 1})"
                if region["has_header"]:
                    header_row_idx = region["min_row"]
                    header_values = []
                    for col_idx in range(region["min_col"], region["max_col"] + 1):
                        cell = sheet.cell(row=header_row_idx, column=col_idx)
                        if cell.value is not None:
                            header_values.append(str(cell.value))

                    if header_values:
                        preview += f" with headers: {', '.join(header_values[:3])}"
                        if len(header_values) > 3:
                            preview += ", ..."

                # Generate a textual representation of the table for searching
                table_content = []
                for row_idx in range(region["min_row"], region["max_row"] + 1):
                    row_values = []
                    for col_idx in range(region["min_col"], region["max_col"] + 1):
                        cell = sheet.cell(row=row_idx, column=col_idx)
                        row_values.append(str(cell.value) if cell.value is not None else "")
                    table_content.append("\t".join(row_values))

                # Join all rows with newlines to create a searchable text representation
                table_text = "\n".join(table_content)

                # Get range in Excel notation (e.g., "A1:F20")
                from openpyxl.utils import get_column_letter
                min_col_letter = get_column_letter(region["min_col"])
                max_col_letter = get_column_letter(region["max_col"])
                range_str = f"{min_col_letter}{region['min_row']}:{max_col_letter}{region['max_row']}"

                # Create data table element
                table_element = {
                    "element_id": table_id,
                    "doc_id": doc_id,
                    "element_type": "data_table",
                    "parent_id": data_tables_id,
                    "content_preview": preview,
                    "content_location": json.dumps({
                        "source": source_id,
                        "type": "data_table",
                        "sheet_name": sheet.title,
                        "range": range_str
                    }),
                    "content_hash": self._generate_hash(table_text),
                    "metadata": {
                        "range": range_str,
                        "min_row": region["min_row"],
                        "min_col": region["min_col"],
                        "max_row": region["max_row"],
                        "max_col": region["max_col"],
                        "row_count": region["max_row"] - region["min_row"] + 1,
                        "column_count": region["max_col"] - region["min_col"] + 1,
                        "has_header": region["has_header"],
                        "detection_confidence": region["confidence"],
                        "sheet": sheet.title,
                        "table_contents": table_text  # Include text version for searching
                    }
                }

                elements.append(table_element)

                # Create relationship from data tables container to data table
                contains_table_relationship = {
                    "relationship_id": self._generate_id("rel_"),
                    "source_id": data_tables_id,
                    "target_id": table_id,
                    "relationship_type": RelationshipType.CONTAINS.value,
                    "metadata": {
                        "confidence": 1.0,
                        "index": idx
                    }
                }
                relationships.append(contains_table_relationship)

                # Create inverse relationship
                table_contained_relationship = {
                    "relationship_id": self._generate_id("rel_"),
                    "source_id": table_id,
                    "target_id": data_tables_id,
                    "relationship_type": RelationshipType.CONTAINED_BY.value,
                    "metadata": {
                        "confidence": 1.0
                    }
                }
                relationships.append(table_contained_relationship)

                # Add separate elements for header row and first column if they're likely headers
                if region["has_header"]:
                    # Extract header row
                    header_id = self._generate_id(f"table_header_row_{idx + 1}_")
                    header_values = []
                    for col_idx in range(region["min_col"], region["max_col"] + 1):
                        cell = sheet.cell(row=region["min_row"], column=col_idx)
                        header_values.append(str(cell.value) if cell.value is not None else "")

                    header_text = "\t".join(header_values)

                    # Create header row element
                    header_element = {
                        "element_id": header_id,
                        "doc_id": doc_id,
                        "element_type": "table_header_row",
                        "parent_id": table_id,
                        "content_preview": header_text[:self.max_content_preview] + (
                            "..." if len(header_text) > self.max_content_preview else ""),
                        "content_location": json.dumps({
                            "source": source_id,
                            "type": "table_header_row",
                            "sheet_name": sheet.title,
                            "range": f"{min_col_letter}{region['min_row']}:{max_col_letter}{region['min_row']}"
                        }),
                        "content_hash": self._generate_hash(header_text),
                        "metadata": {
                            "row": region["min_row"],
                            "values": header_values,
                            "sheet": sheet.title
                        }
                    }

                    elements.append(header_element)

                    # Create relationship from table to header row
                    contains_header_row_relationship = {
                        "relationship_id": self._generate_id("rel_"),
                        "source_id": table_id,
                        "target_id": header_id,
                        "relationship_type": RelationshipType.CONTAINS_TABLE_ROW.value,
                        "metadata": {
                            "confidence": 1.0,
                            "is_header": True
                        }
                    }
                    relationships.append(contains_header_row_relationship)

                    # Create inverse relationship
                    header_row_contained_relationship = {
                        "relationship_id": self._generate_id("rel_"),
                        "source_id": header_id,
                        "target_id": table_id,
                        "relationship_type": RelationshipType.CONTAINED_BY.value,
                        "metadata": {
                            "confidence": 1.0
                        }
                    }
                    relationships.append(header_row_contained_relationship)

                    # Create header cells
                    for col_idx in range(region["min_col"], region["max_col"] + 1):
                        col_pos = col_idx - region["min_col"]
                        cell = sheet.cell(row=region["min_row"], column=col_idx)
                        cell_value = str(cell.value) if cell.value is not None else ""

                        # Only create cell elements for non-empty cells
                        if cell_value:
                            header_cell_id = self._generate_id(f"table_header_cell_{idx + 1}_{col_pos}_")

                            # Create header cell element
                            header_cell_element = {
                                "element_id": header_cell_id,
                                "doc_id": doc_id,
                                "element_type": "table_header",
                                "parent_id": header_id,
                                "content_preview": cell_value[:self.max_content_preview] + (
                                    "..." if len(cell_value) > self.max_content_preview else ""),
                                "content_location": json.dumps({
                                    "source": source_id,
                                    "type": "table_header",
                                    "sheet_name": sheet.title,
                                    "cell": f"{get_column_letter(col_idx)}{region['min_row']}"
                                }),
                                "content_hash": self._generate_hash(cell_value),
                                "metadata": {
                                    "row": region["min_row"],
                                    "column": col_idx,
                                    "value": cell_value,
                                    "sheet": sheet.title
                                }
                            }

                            elements.append(header_cell_element)

                            # Create relationship from header row to header cell
                            contains_header_cell_relationship = {
                                "relationship_id": self._generate_id("rel_"),
                                "source_id": header_id,
                                "target_id": header_cell_id,
                                "relationship_type": RelationshipType.CONTAINS_TABLE_HEADER.value,
                                "metadata": {
                                    "confidence": 1.0,
                                    "col_index": col_pos
                                }
                            }
                            relationships.append(contains_header_cell_relationship)

                            # Create inverse relationship
                            header_cell_contained_relationship = {
                                "relationship_id": self._generate_id("rel_"),
                                "source_id": header_cell_id,
                                "target_id": header_id,
                                "relationship_type": RelationshipType.CONTAINED_BY.value,
                                "metadata": {
                                    "confidence": 1.0
                                }
                            }
                            relationships.append(header_cell_contained_relationship)

                    # Check if first column might contain row headers
                    # Simple heuristic: Check if formatting of first column is different
                    first_col_headers = []
                    first_col_formatting_count = 0

                    for row_idx in range(region["min_row"] + 1, region["max_row"] + 1):  # Skip header row
                        cell = sheet.cell(row=row_idx, column=region["min_col"])
                        first_col_headers.append(str(cell.value) if cell.value is not None else "")

                        # Check if the cell has special formatting (bold, fill, etc.)
                        if (hasattr(cell, 'font') and cell.font and hasattr(cell.font, 'bold') and cell.font.bold) or \
                                (hasattr(cell, 'fill') and cell.fill and hasattr(cell.fill,
                                                                                 'fill_type') and cell.fill.fill_type != 'none'):
                            first_col_formatting_count += 1

                    # If more than 1/3 of cells in the first column have special formatting, consider it a header column
                    if first_col_formatting_count > (region["max_row"] - region["min_row"]) / 3:
                        col_header_id = self._generate_id(f"table_row_headers_{idx + 1}_")
                        col_header_text = "\n".join(first_col_headers)

                        # Create row headers element
                        col_header_element = {
                            "element_id": col_header_id,
                            "doc_id": doc_id,
                            "element_type": "table_row_headers",
                            "parent_id": table_id,
                            "content_preview": col_header_text[:self.max_content_preview] + (
                                "..." if len(col_header_text) > self.max_content_preview else ""),
                            "content_location": json.dumps({
                                "source": source_id,
                                "type": "table_row_headers",
                                "sheet_name": sheet.title,
                                "range": f"{min_col_letter}{region['min_row'] + 1}:{min_col_letter}{region['max_row']}"
                            }),
                            "content_hash": self._generate_hash(col_header_text),
                            "metadata": {
                                "column": region["min_col"],
                                "values": first_col_headers,
                                "sheet": sheet.title
                            }
                        }

                        elements.append(col_header_element)

                        # Create relationship from table to row headers
                        contains_col_header_relationship = {
                            "relationship_id": self._generate_id("rel_"),
                            "source_id": table_id,
                            "target_id": col_header_id,
                            "relationship_type": RelationshipType.CONTAINS.value,
                            "metadata": {
                                "confidence": 1.0,
                                "is_row_headers": True
                            }
                        }
                        relationships.append(contains_col_header_relationship)

                        # Create inverse relationship
                        col_header_contained_relationship = {
                            "relationship_id": self._generate_id("rel_"),
                            "source_id": col_header_id,
                            "target_id": table_id,
                            "relationship_type": RelationshipType.CONTAINED_BY.value,
                            "metadata": {
                                "confidence": 1.0
                            }
                        }
                        relationships.append(col_header_contained_relationship)

        return elements, relationships

    @staticmethod
    def _extract_workbook_links(workbook: openpyxl.workbook.Workbook, elements: List[Dict[str, Any]]) -> List[
        Dict[str, Any]]:
        """
        Extract hyperlinks from Excel workbook elements.
        This is called during the parsing phase to identify hyperlinks in cells.

        Args:
            workbook: The Excel workbook object
            elements: List of extracted elements

        Returns:
            List of hyperlink dictionaries
        """
        links = []

        try:
            # Build a mapping of elements by ID for efficient lookup
            element_map = {element["element_id"]: element for element in elements}

            # Create a mapping from cell references to element IDs
            cell_to_element = {}
            for element in elements:
                if element["element_type"] in ["table_cell", "table_header"]:
                    metadata = element.get("metadata", {})
                    if "address" in metadata:
                        sheet_name = metadata.get("sheet", "")
                        cell_address = metadata["address"]
                        # Use composite key of sheet + cell address for uniqueness
                        cell_to_element[(sheet_name, cell_address)] = element["element_id"]

            # Check each sheet for hyperlinks
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]

                # Skip if sheet doesn't have hyperlinks or appropriate properties
                if not hasattr(sheet, "_hyperlinks") or not sheet._hyperlinks:
                    continue

                # Process each hyperlink in the sheet
                for hyperlink in sheet._hyperlinks:
                    if not hyperlink or not hasattr(hyperlink, "target") or not hyperlink.target:
                        continue

                    # Get cell reference
                    cell_ref = hyperlink.ref if hasattr(hyperlink, "ref") else None
                    if not cell_ref:
                        continue

                    # Find the element ID for this cell
                    element_id = cell_to_element.get((sheet_name, cell_ref))
                    if not element_id:
                        continue

                    # Get the element for display text
                    element = element_map.get(element_id)
                    if not element:
                        continue

                    # Get the display text (either from hyperlink or cell value)
                    display_text = ""
                    if hasattr(hyperlink, "display") and hyperlink.display:
                        display_text = hyperlink.display
                    else:
                        display_text = element.get("content_preview", "")

                    # Get the target URL
                    target = hyperlink.target

                    # Handle relative internal links (e.g., sheet references)
                    if target.startswith("#"):
                        link_type = "internal"
                    else:
                        link_type = "hyperlink"

                    # Create link entry
                    links.append({
                        "source_id": element_id,
                        "link_text": display_text,
                        "link_target": target,
                        "link_type": link_type
                    })

            # Find cells with hyperlinks in the style metadata
            # (For hyperlinks defined via cell formatting rather than hyperlink objects)
            for element in elements:
                if element["element_type"] in ["table_cell", "table_header"] and "metadata" in element:
                    metadata = element["metadata"]

                    # Skip if already processed from _hyperlinks
                    if any(link["source_id"] == element["element_id"] for link in links):
                        continue

                    # Check if hyperlink is in the cell style
                    if "style" in metadata and "hyperlink" in metadata["style"]:
                        hyperlink = metadata["style"]["hyperlink"]

                        # Extract hyperlink details
                        link_target = hyperlink.get("target", "")
                        link_text = hyperlink.get("display", element.get("content_preview", ""))

                        if link_target:
                            links.append({
                                "source_id": element["element_id"],
                                "link_text": link_text,
                                "link_target": link_target,
                                "link_type": "hyperlink"
                            })
        except Exception as e:
            logger.warning(f"Error extracting workbook links: {str(e)}")

        return links

    @staticmethod
    def _is_cell_in_range(cell_coord: str, range_string: str) -> bool:
        """
        Check if a cell coordinate is within a range.

        Args:
            cell_coord: Cell coordinate (e.g., 'A1')
            range_string: Range string (e.g., 'A1:B10')

        Returns:
            True if cell is in range, False otherwise
        """
        try:
            # Use openpyxl's range utility
            from openpyxl.utils.cell import range_boundaries

            # Get cell coordinates
            from openpyxl.utils import column_index_from_string, get_column_letter

            # Parse cell coordinate
            import re
            col_str, row_str = re.match(r'([A-Z]+)([0-9]+)', cell_coord).groups()
            cell_col = column_index_from_string(col_str)
            cell_row = int(row_str)

            # Get range boundaries
            min_col, min_row, max_col, max_row = range_boundaries(range_string)

            # Check if cell is within range
            return (min_col <= cell_col <= max_col) and (min_row <= cell_row <= max_row)
        except Exception:
            # If any error occurs, assume cell is not in range
            return False
