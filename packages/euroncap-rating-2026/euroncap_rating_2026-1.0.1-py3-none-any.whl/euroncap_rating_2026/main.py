# Copyright 2025, Euro NCAP IVZW
# Created by IVEX NV (https://ivex.ai)
#
# Licensed under the Apache License 2.0.
# See http://www.apache.org/licenses/LICENSE-2.0 for details.

import pandas as pd
import logging

from euroncap_rating_2026.data_loader import load_data, compute_score
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
import sys
from pandas.api.types import is_string_dtype
from pydantic_settings import BaseSettings, SettingsConfigDict
from pydantic import Field
import argparse
import os
from datetime import datetime
from openpyxl.styles import Alignment
from openpyxl.styles import Border, Side
import shutil
from importlib.resources import files


class Settings(BaseSettings):
    log_level: str = Field(
        default="INFO",
        description="Logging level for the application (DEBUG, INFO, WARNING, ERROR, CRITICAL).",
    )
    enable_debug_gui: bool = Field(
        default=False,
        description="Enable debug mode for detailed logging.",
    )
    model_config = SettingsConfigDict(env_prefix="euroncap_rating_2026_")


settings = Settings()


def logging_config(output_path: str):
    # Ensure the output directory exists
    os.makedirs(output_path, exist_ok=True)

    # Configure logging
    logging.basicConfig(
        level=getattr(
            logging, settings.log_level.upper(), logging.INFO
        ),  # Use settings log level
        format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
        handlers=[
            logging.FileHandler(
                os.path.join(output_path, "euroncap_rating_2026.log"), mode="w"
            ),  # Log to file
        ],
    )

    logger = logging.getLogger()  # Root logger
    logger.info("Logging is set up!")


logger = logging.getLogger(__name__)

if settings.enable_debug_gui:
    try:
        from euroncap_rating_2026.debug_gui import show_criteria
    except ImportError:
        print("tkinter is not available. Please install it to use the GUI features.")
        print("To use the GUI features, please install the required dependencies:")
        print("Windows: Install Python's Tkinter module via the Python installer.")
        print(
            "Mac: Tkinter is included with Python on macOS. Ensure you have Python installed."
        )
        print("Linux: Run 'sudo apt-get install python3-tk' to install Tkinter.")
        sys.exit(1)


def generate_template():
    """
    Copies the template.xlsx file to the current working directory.
    """
    template_path = str(files("data").joinpath("template.xlsx"))
    dest_path = os.path.join(os.getcwd(), "template.xlsx")
    shutil.copyfile(template_path, dest_path)
    print(f"Template generated at {dest_path}")


def get_current_loadcase_id(df, index):
    current_loadcase_id = df.loc[index, "Loadcase"]
    if not pd.isna(df.iloc[index]["Seat position"]):
        current_loadcase_id += f"_{df.iloc[index]['Seat position']}"
    if not pd.isna(df.iloc[index]["Dummy"]):
        current_loadcase_id += f"_{df.iloc[index]['Dummy']}"

    # Then process the remaining rows
    for i in range(index + 1, len(df)):
        next_row = df.iloc[i]
        if not pd.isna(next_row["Loadcase"]):
            break
        if not pd.isna(next_row["Seat position"]):
            current_loadcase_id += f"_{next_row['Seat position']}"
        if not pd.isna(next_row["Dummy"]):
            current_loadcase_id += f"_{next_row['Dummy']}"
    return current_loadcase_id


def update_loadcase(df, loadcase):
    last_load_case = None
    last_seat_name = None
    last_dummy_name = None
    last_body_region = None

    if len(loadcase.raw_seats) > 0:
        logger.debug(
            f"Processing loadcase: {loadcase.name} with {len(loadcase.raw_seats)} raw seats"
        )
        logger.debug(f"raw_seats: {[s.name for s in loadcase.raw_seats]}")
    # Ensure the DataFrame has the "Score" and "Capping?" columns
    if "Score" not in df.columns:
        df["Score"] = ""
    if "Capping?" not in df.columns:
        df["Capping?"] = ""

    for column in ["Colour", "Capping?", "Prediction.Check"]:
        if column in df.columns and not is_string_dtype(df[column]):
            df[column] = df[column].astype(str)

    for index, row in df.iterrows():
        if not pd.isna(row["Loadcase"]):
            last_load_case = row["Loadcase"]
            current_loadcase_id = get_current_loadcase_id(df, index)
        if not pd.isna(row["Seat position"]):
            last_seat_name = row["Seat position"]
        if not pd.isna(row["Dummy"]):
            last_dummy_name = row["Dummy"]
        if not pd.isna(row["Body Region"]):
            last_body_region = row["Body Region"]
        if (
            not (
                "Static-Front" in current_loadcase_id and "Static-Front" in loadcase.id
            )
            and current_loadcase_id != loadcase.id
        ):
            continue

        criteria = row["Criteria"]
        # Use loadcase.raw_seats if available, otherwise use loadcase.seats
        seat_list = (
            loadcase.raw_seats
            if hasattr(loadcase, "raw_seats") and len(loadcase.raw_seats) > 0
            else loadcase.seats
        )
        logger.debug(f"seat_list: {[f'{s.name} ({s.dummy.name})' for s in seat_list]}")
        seat = next(
            (
                s
                for s in seat_list
                if s.name == last_seat_name and s.dummy.name == last_dummy_name
            ),
            None,
        )
        if seat is None:
            logger.debug(
                f"Seat {last_seat_name} with dummy {last_dummy_name} not found in loadcase {loadcase.name}"
            )
            return

        body_region = next(
            (br for br in seat.dummy.body_region_list if br.name == last_body_region),
            None,
        )
        if body_region is None:
            logger.debug(
                f"Body region {last_body_region} not found in dummy {seat.dummy.name}"
            )

        criteria_obj = next(
            (c for c in body_region._criteria if c.name == criteria), None
        )
        logger.debug(f"Loadcase: {loadcase.id}, Criteria object: {criteria_obj}")
        if criteria_obj:
            df.loc[index, "HPL"] = criteria_obj.hpl
            df.loc[index, "LPL"] = criteria_obj.lpl
            df.loc[index, "Colour"] = criteria_obj.color
            df.loc[index, "Score"] = criteria_obj.score
            df.loc[index, "Capping?"] = "YES" if criteria_obj.capping else ""
            if criteria_obj.prediction_result and "Prediction.Check" in df.columns:
                df.loc[index, "Prediction.Check"] = "".join(
                    word.capitalize() for word in criteria_obj.prediction_result.split()
                )
            logger.debug(
                f"Updated row - Loadcase: {current_loadcase_id}, Seat: {last_seat_name}, Dummy: {last_dummy_name}, Body Region: {last_body_region}, Criteria: {criteria}, Colour: {criteria_obj.color}, Score: {criteria_obj.score}, Capping: {criteria_obj.capping}, Prediction: {criteria_obj.prediction_result}"
            )
    return df


def update_dummy_scores(df, loadcase):

    last_load_case = None
    last_seat_name = None
    last_dummy_name = None

    for column in ["Capping?"]:
        if column in df.columns and not is_string_dtype(df[column]):
            df[column] = df[column].astype(str)

    for index, row in df.iterrows():
        if not pd.isna(row["Loadcase"]):
            last_load_case = row["Loadcase"]
            current_loadcase_id = get_current_loadcase_id(df, index)
        if not pd.isna(row["Seat position"]):
            last_seat_name = row["Seat position"]
        if not pd.isna(row["Dummy"]):
            last_dummy_name = row["Dummy"]

        if current_loadcase_id != loadcase.id:
            continue

        seat = next(
            (
                s
                for s in loadcase.seats
                if s.name == last_seat_name and s.dummy.name == last_dummy_name
            ),
            None,
        )
        if seat is None:
            logger.debug(
                f"Seat {last_seat_name} with dummy {last_dummy_name} not found in loadcase {loadcase.name}"
            )
            return

        if seat.dummy.score is not None:
            df.loc[index, "Score"] = seat.dummy.score
        df.loc[index, "Capping?"] = "Capped" if seat.dummy.capping else ""

        logger.debug(
            f"Updated row - Loadcase: {last_load_case}, Seat: {last_seat_name}, Dummy: {last_dummy_name}, Score: {seat.dummy.score}"
        )

    return df


def update_bodyregion(df, loadcase):

    last_load_case = None
    last_seat_name = None
    last_dummy_name = None
    last_body_region = None

    for index, row in df.iterrows():
        if not pd.isna(row["Loadcase"]):
            last_load_case = row["Loadcase"]
            current_loadcase_id = get_current_loadcase_id(df, index)
        if not pd.isna(row["Seat position"]):
            last_seat_name = row["Seat position"]
        if not pd.isna(row["Dummy"]):
            last_dummy_name = row["Dummy"]
        if not pd.isna(row["Body Region"]):
            last_body_region = row["Body Region"]

        if current_loadcase_id != loadcase.id:
            continue

        seat = next(
            (
                s
                for s in loadcase.seats
                if s.name == last_seat_name and s.dummy.name == last_dummy_name
            ),
            None,
        )
        if seat is None:
            logger.debug(
                f"Seat {last_seat_name} with dummy {last_dummy_name} not found in loadcase {loadcase.name}"
            )
            return

        body_region = next(
            (br for br in seat.dummy.body_region_list if br.name == last_body_region),
            None,
        )
        if body_region is None:
            logger.debug(
                f"Body region {last_body_region} not found in dummy {seat.dummy.name}"
            )
            return

        df.loc[index, "Body regionscore"] = body_region.bodyregion_score
        df.loc[index, "Score"] = body_region.score
        df.loc[index, "Modifiers"] = sum(
            measurement.modifier
            for measurement in body_region._measurement
            if measurement.modifier is not None
        )

        logger.debug(
            f"Updated row - Loadcase: {last_load_case}, Seat: {last_seat_name}, Dummy: {last_dummy_name}, Body Region: {last_body_region}, Score: {body_region.score}"
        )
    return df


def update_stage_scores(df, final_scores):

    last_stage_subelement = None

    for index, row in df.iterrows():

        if not pd.isna(row["Stage Subelement"]):
            last_stage_subelement = row["Stage Subelement"]

        for key in final_scores:
            if key == last_stage_subelement:
                df.loc[index, "Score"] = final_scores[key]

    return df


def is_empty_cell(cell_value):
    """
    Checks if a cell value is considered empty.

    Args:
        cell_value: The value of the cell to check.

    Returns:
        bool: True if the cell is empty, False otherwise.
    """
    return (
        pd.isna(cell_value)
        or cell_value is None
        or cell_value == "None"
        or cell_value == "nan"
    )


def args_entrypoint() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        prog="euroncap_rating_2026",
        usage="%(prog)s <command> [options]",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
        epilog="%(prog)s -h for help",
        add_help=True,
        allow_abbrev=False,
        description="Euro NCAP Rating Calculator 2026 application to compute NCAP scores.",
    )

    subparsers = parser.add_subparsers(
        dest="command", required=True, help="Sub-commands"
    )

    # compute command
    compute_parser = subparsers.add_parser(
        "compute",
        help="Compute NCAP scores from an input Excel file.",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    compute_parser.add_argument(
        "--input_file",
        type=str,
        required=True,
        help="Path to the input Excel file containing NCAP test measurements.",
    )
    compute_parser.add_argument(
        "--output_path",
        type=str,
        default=os.getcwd(),
        help="Path to the output directory where the report will be saved.",
    )

    # generate_template command
    template_parser = subparsers.add_parser(
        "generate_template",
        help="Generate template.xlsx file to the current working directory.",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )

    return parser.parse_args()


def main():
    """
    Main function to load data, compute scores, and display the criteria values.
    """
    args = args_entrypoint()
    if args.command == "generate_template":
        generate_template()
        print("Template generated successfully.")
        return
    elif args.command != "compute":
        print(f"Unknown command: {args.command}")
        sys.exit(1)
    input_file = args.input_file
    output_path = args.output_path

    logger.info("Start main")
    logging_config(output_path)

    if not input_file:
        logger.error("Input file path is required.")
        sys.exit(1)
    if not input_file.endswith(".xlsx"):
        logger.error("Input file must be an Excel file with .xlsx extension.")
        sys.exit(1)

    print("-" * 40)
    print("Run Settings")
    print("-" * 40)
    print(f"Log Level: {settings.log_level}")
    logger.info(f"Log Level: {settings.log_level}")
    print(f"Enable Debug GUI: {settings.enable_debug_gui}")
    logger.info(f"Enable Debug GUI: {settings.enable_debug_gui}")
    print(f"Output report path: {output_path}")
    logger.info(f"Output report path: {output_path}")
    print("-" * 40)

    print("Loading data from spreadsheet...")
    sheet_dict, test_score_inspection = load_data(input_file)

    if settings.enable_debug_gui:
        show_criteria(sheet_dict)

    print("Computing NCAP scores...")
    overall_score, overall_max_score, final_scores, final_max_scores = compute_score(
        sheet_dict, test_score_inspection
    )

    current_datetime = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    output_file = os.path.join(output_path, f"{current_datetime}_report.xlsx")

    # Copy specific sheets to the output file
    sheets_to_copy = [
        "Test Scores",
        "CP - Dummy Scores",
        "CP - Body Region Scores",
        "Input parameters",
    ]
    score_df_dict = {}
    for i, sheet in enumerate(sheets_to_copy):
        try:
            if i == 0:
                mode = "w"
            else:
                mode = "a"
            df = pd.read_excel(input_file, sheet_name=sheet, header=0)
            score_df_dict[sheet] = df  # Store the DataFrame for later use
            with pd.ExcelWriter(output_file, engine="openpyxl", mode=mode) as writer:
                df.to_excel(writer, sheet_name=sheet, index=False)
        except Exception as e:
            logger.error(f"Failed to copy sheet {sheet}: {e}")

    for sheet_name in sheet_dict:

        # Copy the sheet name from the input file to the output file
        writer = pd.ExcelWriter(output_file, engine="openpyxl", mode="a")
        df = pd.read_excel(input_file, sheet_name=sheet_name, header=0)
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        writer.close()
        logger.info(f"score_df_dict keys: {list(score_df_dict.keys())}")
        for i, loadcase_df in enumerate(sheet_dict[sheet_name]):
            df = update_loadcase(df, loadcase_df)
            logger.debug(f"Processing loadcase_df: {loadcase_df}")
            score_df_dict["CP - Body Region Scores"] = update_bodyregion(
                score_df_dict["CP - Body Region Scores"], loadcase_df
            )
            score_df_dict["CP - Dummy Scores"] = update_dummy_scores(
                score_df_dict["CP - Dummy Scores"], loadcase_df
            )

        # Save the updated DataFrame back to the output file
        writer = pd.ExcelWriter(
            output_file, engine="openpyxl", mode="a", if_sheet_exists="replace"
        )
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        writer.close()

    score_df_dict["Test Scores"] = update_stage_scores(
        score_df_dict["Test Scores"], final_scores
    )

    header_align_right = [
        "Body regionscore",
        "Modifiers",
        "Inspection [%]",
        "Score",
        "Max Score",
        "HPL",
        "LPL",
        "Capping",
        "OEM.Prediction",
        "Value",
        "Prediction.Check",
        "Colour",
        "Points",
        "Modifier",
    ]
    header_align_left = [
        "Loadcase",
        "Seat position",
        "Body Region",
        "Criteria",
        "Stage",
        "Stage element",
        "Stage subelement",
        "Dummy",
    ]
    # Write the updated copied sheets back to the output file
    for sheet_name, updated_df in score_df_dict.items():
        try:
            with pd.ExcelWriter(
                output_file, engine="openpyxl", mode="a", if_sheet_exists="replace"
            ) as writer:
                updated_df.to_excel(writer, sheet_name=sheet_name, index=False)
        except Exception as e:
            logger.error(f"Failed to write updated sheet {sheet_name}: {e}")

    try:
        wb = load_workbook(output_file)

        calculation_color = PatternFill(
            start_color="FFDD04", end_color="FFDD04", fill_type="solid"
        )
        input_color = PatternFill(
            start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"
        )

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            ###############################################################
            # Set first row (header) to black background and white text
            ###############################################################
            header_fill = PatternFill(
                start_color="000000", end_color="000000", fill_type="solid"
            )
            header_font = Font(name="Calibri", color="FFFFFF", bold=True)
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font

            ###############################################################
            # Align text
            ###############################################################
            for col in ws.iter_cols():
                header = col[0].value
                if header in header_align_right:
                    align = Alignment(horizontal="right")
                elif header in header_align_left:
                    align = Alignment(horizontal="left")
                for cell in col:
                    cell.alignment = align

                if col[0].value in ["Colour", "OEM.Prediction"]:
                    for cell in col[1:]:
                        if (
                            isinstance(cell.value, str)
                            and cell.value
                            and str(cell.value).lower() not in ["nan", "none"]
                        ):
                            cell.value = cell.value.capitalize()

            ###############################################################
            # Set number precision for specific columns
            ###############################################################
            for col in ws.iter_cols():
                header = col[0].value
                if header in [
                    "Inspection [%]",
                    "Value",
                    "Capping",
                    "Body regionscore",
                    "Modifiers",
                ]:
                    for cell in col[1:]:
                        if isinstance(cell.value, (int, float)) and not is_empty_cell(
                            cell.value
                        ):
                            cell.number_format = "0.00"
                elif header in ["Score", "Max Score", "Max score"]:
                    for cell in col[1:]:
                        if isinstance(cell.value, (int, float)) and not is_empty_cell(
                            cell.value
                        ):
                            if ws.title not in score_df_dict.keys():
                                cell.number_format = "0.00"
                            elif ws.title == "Test Scores":
                                cell.number_format = "0.000"
                            else:
                                cell.number_format = "0.0000"

            ###############################################################
            # Color columns based on their headers
            ###############################################################
            for col in ws.iter_cols():
                if col[0].value in [
                    "Score",
                    "Points",
                    "Prediction.Check",
                    "Body regionscore",
                    "Modifiers",
                    "Modifier",
                    "Colour",
                    "Capping?",
                ]:
                    for cell in col[1:]:  # Skip the first row (header)
                        cell.fill = calculation_color
                        if is_empty_cell(cell.value):
                            cell.value = ""
                elif col[0].value in ["Inspection [%]", "Value", "OEM.Prediction"]:
                    for cell in col[1:]:  # Skip the first row (header)
                        cell.fill = input_color
                        if is_empty_cell(cell.value):
                            cell.value = ""

        ###############################################################
        # Adjust column widths based on content
        ###############################################################
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter  # Get the column letter
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except (TypeError, AttributeError) as e:
                        logger.warning(
                            f"Error processing cell value '{cell.value}': {e}"
                        )
                adjusted_width = max_length + 2  # Add some padding
                ws.column_dimensions[column_letter].width = adjusted_width

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            prev_value = None
            # Add thick border to the last column for all rows including header
            for row in range(1, ws.max_row + 1):  # Start from 1 to include header
                cell = ws.cell(row=row, column=ws.max_column)
                cell.border = cell.border + Border(
                    right=Side(border_style="thick", color="000000")
                )
            # Add thick border to the last row for all columns
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=ws.max_row, column=col)
                cell.border = cell.border + Border(
                    bottom=Side(border_style="thick", color="000000")
                )
            for row in range(1, ws.max_row + 1):  # Skip header row
                cell = ws.cell(row=row, column=1)
                # Set top border when previous value is empty (nan/None) and current is not empty
                if (
                    (
                        prev_value is None
                        or str(prev_value).lower() in ["nan", "none", ""]
                    )
                    and (
                        cell.value is not None
                        and str(cell.value).lower() not in ["nan", "none", ""]
                    )
                    and row != 2
                ):
                    for col in range(1, ws.max_column + 1):
                        ws.cell(row=row, column=col).border = ws.cell(
                            row=row, column=col
                        ).border + Border(
                            top=Side(border_style="thick", color="000000")
                        )
                prev_value = cell.value
        wb.save(output_file)
    except Exception as e:
        logger.error(f"Failed to apply formatting: {e}")

    print("-" * 40)
    print("Score:")
    print("-" * 40)
    print(
        f"{'Stage Element':<20}{'Stage Subelement':<20}{'Score':<10}{'Max Score':<10}"
    )
    print("-" * 40)
    score_order = [
        ("Frontal Impact", ["Offset", "FW", "Sled & VT"]),
        ("Side Impact", ["MDB", "Pole", "Farside"]),
        ("Rear Impact", ["Whiplash"]),
    ]

    printed_categories = set()
    for category, subcategories in score_order:
        for subcategory in subcategories:
            if subcategory in final_scores:
                logger.info(
                    f"Final score for {subcategory}: {final_scores[subcategory]}/{final_max_scores[subcategory]}"
                )
                category_to_print = (
                    category if category not in printed_categories else ""
                )
                print(
                    f"{category_to_print:<20}{subcategory:<20}{final_scores[subcategory]:<10}{final_max_scores[subcategory]:<10}"
                )
                printed_categories.add(category)
            else:
                logger.warning(f"Score for {subcategory} not found.")
                category_to_print = (
                    category if category not in printed_categories else ""
                )
                print(f"{category_to_print:<20}{subcategory:<20}{'N/A':<10}{'N/A':<10}")
                printed_categories.add(category)

    print("-" * 40)
    print(" " * 40)

    overall_str = "Final score"
    print(f"{overall_str:<20}{overall_score:<10}{overall_max_score:<10}")
    print(" " * 40)
    print(f"Log available at {os.path.join(output_path, 'euroncap_rating_2026.log')}")
    print(" " * 40)
    print(f"Final report available at {output_file}")
    logger.info(f"Final report available at {output_file}")


if __name__ == "__main__":
    main()
