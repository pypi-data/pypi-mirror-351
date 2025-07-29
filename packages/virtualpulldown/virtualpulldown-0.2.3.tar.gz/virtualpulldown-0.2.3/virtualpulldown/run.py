#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import os, sys
from os.path import join
import re
import subprocess
import glob
from Bio import SeqIO
from openpyxl import Workbook
import glob
from PIL import Image, ImageDraw, ImageFont
import json
import numpy as np

import os
import sys

def pdb_to_fasta(path, pdb_file, output_name):
    aa = {
        "GLY": "G",
        "ALA": "A",
        "SER": "S",
        "PRO": "P",
        "VAL": "V",
        "THR": "T",
        "CYS": "C",
        "LEU": "L",
        "ILE": "I",
        "ASN": "N",
        "ASP": "D",
        "GLN": "Q",
        "LYS": "K",
        "GLU": "E",
        "MET": "M",
        "HIS": "H",
        "PHE": "F",
        "ARG": "R",
        "TYR": "Y",
        "TRP": "W",
        "MSE": "M",
        "UNK": "X"
    }
    resn = []
    seq = ""
    txt_path = os.path.join(path, pdb_file)
    with open(txt_path) as file:
        for line in file:
            try:
                c1, c2, c3, c4 = line.split()[:4]
                if c3 == "CA":
                    resn.append(c4)
            except ValueError:
                pass
    for i in resn:
        seq += aa[i]
    with open(os.path.join(path, f"{output_name}.fasta"), "w") as f:
        f.write(f">{output_name}\n{seq}")

def convert_to_pdf(figures):
    for figure in figures:
        ext=os.path.splitext(figure)[1]
        if ext == '.eps':
            output_file = os.path.splitext(figure)[0]+'.pdf'
            subprocess.call(['ps2pdf','-dEPSCrop', figure, output_file])
        elif ext == '.png':
            output_file = os.path.splitext(figure)[0]+'.pdf'
            subprocess.call(['convert', figure, output_file])
        else:
            raise ValueError('Unsupported file format: ' +ext)

def rename_files(path):
    new_names = []

    # Get a list of files in the directory
    files = os.listdir(path)

    # Filter the files to get only the ones with .a3m extension
    a3m_files = sorted([f for f in files if f.endswith(".a3m")])
    a3m_files = sorted([f for f in a3m_files if f.split(".")[0].isdigit()], key=lambda x: int(x.split(".")[0]))

    # Get the second line of each a3m file and append it to the new_names list
    for a3m_file in a3m_files:
        with open(os.path.join(path, a3m_file), "r") as f:
            lines = f.readlines()
            second_line = lines[1]
            new_names.append(second_line.split()[0])

    # Create a dictionary to map the old filenames to the new names
    name_dict = {}
    for a3m_file, new_file in zip(a3m_files, new_names):
        keyword = a3m_file.split(".")[0] + '_'
        name_dict[keyword] = new_file[1:]

    # Rename the files with the new names
    for file in files:
        # Check if file starts with any of the keys in name_dict
        for key in name_dict.keys():
            if file.startswith(key):
                new_filename = file.replace(key, name_dict[key]+'_', 1)
                # Rename the file
                os.rename(os.path.join(path, file), os.path.join(path, new_filename))
                break

def plot_ptm_iptm(bait_name, title_offset, path, f_width, f_height, fontsize, margin_top, margin_bot, margin_left, margin_right, key_position):
    ptms=[]
    iptms=[]
    pae_data=[]
    gnu_data=[]
    json_files = glob.glob(f'{path}*_seed_000.json')
    for json_file in sorted(json_files):
        with open (json_file) as f:
            data=json.load(f)
            ptms.append(data['ptm'])
            iptms.append(data['iptm'])
    for i, p, ip in zip(sorted(json_files), ptms, iptms):
        pae_data.append(\
    f"{i.split('_scores_rank')[0][2:]+'_'+i.split('_scores_rank_00')[1][:1]} {p:.2f} {ip:.2f}")
    
    for data in pae_data:
        gnu_data.append(data.replace('_','.'))
    
    
    # Plot the graph using gnuplot
    with open('%s.gp'%(bait_name), 'w') as f:
        # Define the plot settings
        f.write('set term xterm\n')
        f.write('set tmargin %d\n'%(margin_top))
        f.write('set bmargin %d\n'%(margin_bot))
        f.write('set lmargin %d\n'%(margin_left))
        f.write('set rmargin %d\n'%(margin_right))
        f.write('set title "%s alphafold pulldown" font "Helvetica-Bold, 18" offset 0,%d \n'%(bait_name,title_offset))
        f.write('set xlabel "predicted models"\n')
        f.write('set ylabel "pTM and ipTM values"\n')
        f.write('set key %s\n'%(key_position))
        f.write('set xtics rotate by -45\n')
        f.write('set key box lt -1 lw 2\n')
        f.write('set x2tics out\n')
        f.write('set x2tics rotate by 45\n')
        f.write('set grid xtics\n')
        f.write('set grid x2tics\n')
        f.write('set terminal postscript eps enhanced color solid "Helvetica" %d size %d,%d\n'%(fontsize, f_width, f_height))
        f.write('set output "%s.eps"\n'%(bait_name))
        # Plot the data
        f.write('plot "-" u 1:3:4:xticlabels(2) w p pt 7 lc rgb "red" notitle, "-" u 1:3:4:x2ticlabel(2) w p pt 7 lc rgb "red" notitle, "-" u 1:4 w lp pt 7 lc rgb "blue" t "ipTM", "-" u 1:3 w lp pt 7 lc rgb "red" t "pTM" \n')
        
        for i in range(0,len(gnu_data),2):
            f.write('{} {}\n'.format(i+1,gnu_data[i]))
        f.write('e\n')
        for i in range(1,len(gnu_data),2): 
            f.write('{} {}\n'.format(i+1,gnu_data[i]))
        f.write('e\n')
        for i in range(len(gnu_data)):
            f.write('{} {}\n'.format(i+1,gnu_data[i]))
        f.write('e\n')
        for i in range(len(gnu_data)):
            f.write('{} {}\n'.format(i+1,gnu_data[i]))
        f.write('e\n')
    # Call gnuplot to create the graph
    subprocess.call(['gnuplot', '%s.gp'%(bait_name)])

def concatenate_images(path):
    # Get all PNG files in the directory
    image_files = glob.glob(f'{path}*pae.png')
        
    # Open all images and their corresponding labels
    images_with_labels = []
    for img_file in sorted(image_files):
        image = Image.open(img_file)
        label = os.path.basename(img_file)
        images_with_labels.append((image, label))
        
    # Get dimensions of the first image
    width, height = images_with_labels[0][0].size
        
    # Create a new image with the same width and the combined height of all images
    result = Image.new('RGB', (width, height * len(images_with_labels)), color='white')
        
    # Paste each image into the result image vertically along with the labels
    title_font = ImageFont.load_default()
    draw = ImageDraw.Draw(result)
    for i, (img, label) in enumerate(images_with_labels):
        result.paste(img, (0, i * height))
        #label_width, label_height = draw.textsize(label, font=title_font)
        draw.text((0, i * height), label, font=title_font, fill=(0, 0, 0), size=46)

    return result

def combine_pairwise_batch(path, filenames, bait_name):
    filenames   = sorted(glob.glob("./fa/"+"*.fa"))
    f2_txt      = f"{bait_name}_bait_truncated.fasta"
    txt2_path   = os.path.join(path,f2_txt)    
    for fa in filenames:
        seq1    = list(SeqIO.parse(fa,'fasta'))
        seq2    = list(SeqIO.parse(txt2_path,'fasta'))
        seq_str = ''
        
        for seq2 in seq2:
            for seq1 in seq1:
                seq_str += '>' + seq1.id + '\n'
                seq_str += str(seq2.seq)+':'+str(seq1.seq) 
                of=open('%s_pair.fasta'%(fa),'w')
                of.write(seq_str)
                of.close()
                seq_str =''

def filter_prey_sequences(path, filename, prey_size_limit, bait_name, filter_start, filter_end):
    f1_txt      = f"{filename}_check.txt"
    f2_txt      = f"{bait_name}.fasta"
    txt_path    = os.path.join(path,f1_txt)
    txt2_path   = os.path.join(path,f2_txt)
    record      = list(SeqIO.parse(txt_path, 'fasta'))
    record2     = list(SeqIO.parse(txt2_path, 'fasta'))

    wb          = Workbook()
    ws          = wb.active
    ws['A1']    = 'Locus_tag'
    ws['B1']    = 'Gene_length'
    ws['C1']    = 'Skip_%s'%(prey_size_limit)
    ws['D1']    = 'start'
    ws['E1']    = 'end'
    ws['F1']    = 'selected'

    for line in record:
        st=line.description.split('[location=')[-1].split('complement(')[-1].split('<')[-1].split('join(')[-1].split(',')[-1].split('..')[0]
        en=line.description.split('[location=')[-1].split('complement(')[-1].split('<')[-1].split('join(')[-1].split(',')[-1].split('..')[-1].split('>')[-1].split(')')[0].split(']')[0]
        if line.seq != record2[0].seq and len(line.seq) < prey_size_limit: 
            if int(st)<=filter_start and int(en) <= filter_start:
                pass
            elif int(st)<=filter_start and int(en) >= filter_start:
                of=open('%s.fa'%(line.description.split('[locus_tag=')[1].split('[protein')[0][:-2]),'w')
                of.write('>'+line.description.split('[locus_tag=')[1].split('[protein')[0][:-2]+'\n')
                of.write(str(line.seq))
                of.close()
            elif int(st) >= filter_start and int(en) <= filter_end:
                of=open('%s.fa'%(line.description.split('[locus_tag=')[1].split('[protein')[0][:-2]),'w')
                of.write('>'+line.description.split('[locus_tag=')[1].split('[protein')[0][:-2]+'\n')
                of.write(str(line.seq))
                of.close()
            elif int(st) >= filter_start and int(st) <= filter_end and int(en) >= filter_end:
                of=open('%s.fa'%(line.description.split('[locus_tag=')[1].split('[protein')[0][:-2]),'w')
                of.write('>'+line.description.split('[locus_tag=')[1].split('[protein')[0][:-2]+'\n')
                of.write(str(line.seq))
                of.close()
            elif int(st) >= filter_end:
                pass
        elif line.seq != record2[0].seq and len(line.seq) > prey_size_limit:
            pass
    for line in record:
        st=line.description.split('[location=')[-1].split('complement(')[-1].split('<')[-1].split('join(')[-1].split(',')[-1].split('..')[0]
        en=line.description.split('[location=')[-1].split('complement(')[-1].split('<')[-1].split('join(')[-1].split(',')[-1].split('..')[-1].split('>')[-1].split(')')[0].split(']')[0]
        if int(st)<= filter_start and int(en) <= filter_start:
            if len(line.seq) < prey_size_limit:
                ws.append([line.description.split('[locus_tag=')[1].split('[protein')[0][:-2]]\
                          +[len(line.seq)]\
                          +['no']\
                          +[st]\
                          +[en]\
                          +['no'])
            else:
                ws.append([line.description.split('[locus_tag=')[1].split('[protein')[0][:-2]]\
                          +[len(line.seq)]\
                          +['yes']\
                          +[st]\
                          +[en]\
                          +['no'])
        elif int(st) <= filter_start and int(en) >= filter_start:
            if len(line.seq) < prey_size_limit:
                ws.append([line.description.split('[locus_tag=')[1].split('[protein')[0][:-2]]\
                          +[len(line.seq)]\
                          +['no']\
                          +[st]\
                          +[en]\
                          +['yes'])
            else:
                ws.append([line.description.split('[locus_tag=')[1].split('[protein')[0][:-2]]\
                          +[len(line.seq)]\
                          +['yes']\
                          +[st]\
                          +[en]\
                          +['size_limit'])
        elif int(st) >= filter_start and int(en) <=filter_end:
            if len(line.seq) < prey_size_limit:
                ws.append([line.description.split('[locus_tag=')[1].split('[protein')[0][:-2]]\
                          +[len(line.seq)]\
                          +['no']\
                          +[st]\
                          +[en]\
                          +['yes'])
            else:
                ws.append([line.description.split('[locus_tag=')[1].split('[protein')[0][:-2]]\
                          +[len(line.seq)]\
                          +['yes']\
                          +[st]\
                          +[en]\
                          +['size_limit'])
        elif int(st) >= filter_start and int(st) <=filter_end and int(en) >=filter_end:
            if len(line.seq) < prey_size_limit:
                ws.append([line.description.split('[locus_tag=')[1].split('[protein')[0][:-2]]\
                          +[len(line.seq)]\
                          +['no']\
                          +[st]\
                          +[en]\
                          +['yes'])
            else:
                ws.append([line.description.split('[locus_tag=')[1].split('[protein')[0][:-2]]\
                          +[len(line.seq)]\
                          +['yes']\
                          +[st]\
                          +[en]\
                          +['size_limit'])
        elif int(line.description.split('..')[0].split("=")[-1].split('(')[-1].split('<')[-1]) >= filter_end:
            if len(line.seq) < prey_size_limit:
                ws.append([line.description.split('[locus_tag=')[1].split('[protein')[0][:-2]]\
                          +[len(line.seq)]\
                          +['no']\
                          +[st]\
                          +[en]\
                          +['no'])
            else:
                ws.append([line.description.split('[locus_tag=')[1].split('[protein')[0][:-2]]\
                          +[len(line.seq)]\
                          +['yes']\
                          +[st]\
                          +[en]\
                          +['no'])
    wb.save("%s%s_filtered.xlsx"%(path,bait_name))

def create_mastertable(path, filename, prey_size_limit, bait_name):
    f1_txt    = f"{filename}_check.txt"
    f2_txt    = f"{bait_name}.fasta"
    txt_path  = os.path.join(path,f1_txt)
    txt2_path = os.path.join(path,f2_txt)
    record    = list(SeqIO.parse(txt_path, 'fasta'))
    record2   = list(SeqIO.parse(txt2_path, 'fasta'))

    wb        = Workbook()
    ws        = wb.active
    ws['A1']  = 'Locus_tag'
    ws['B1']  = 'Gene_description'
    ws['C1']  = 'Gene_length'
    ws['D1']  = 'Skip_%s'%(prey_size_limit)
    ws['E1']  = 'start'
    ws['F1']  = 'end'
    ws['G1']  = '%s_info'%(bait_name)

    for line in record:
        if len(line.seq) <= prey_size_limit:
            ws.append([line.description.split('[locus_tag=')[1].split('[protein')[0][:-2]]\
                      +[line.description.split('protein=')[1].split(']')[0]]\
                      +[len(line.seq)]\
                      +['no']\
                      +[line.description.split('..')[0].split("=")[-1].split('(')[-1].split('<')[-1]]\
                      +[line.description.split('..')[1].split("]")[0].split(')')[0].split('>')[-1]])
        else:
            ws.append([line.description.split('[locus_tag=')[1].split('[protein')[0][:-2]]\
                      +[line.description.split('protein=')[1].split(']')[0]]\
                      +[len(line.seq)]\
                      +['yes']\
                      +[line.description.split('..')[0].split("=")[-1].split('(')[-1].split('<')[-1]]\
                      +[line.description.split('..')[1].split("]")[0].split(')')[0].split('>')[-1]])
    for line in record:
        if line.seq ==record2[0].seq:
            ws['H1']=str(line.description.split('[locus_tag=')[1].split('[protein')[0][:-2])
    wb.save(f"{bait_name}_master.xlsx")

def identify_prophage_region(accession_number):
    url = f'http://phaster.ca/phaster_api?acc={accession_number}'
    result = subprocess.run(['wget', '-qO-', url], stdout=subprocess.PIPE)
    output = result.stdout.decode()
    keywords = re.findall(r'intact\(\d{1,10}\)|questionable\(\d{1,10}\)|incomplete\(\d{1,10}\)', output)
    ranges = re.findall(r'\d{1,10}-\d{1,10}', output)
    if not ranges:
        print("Prophage is not found")
    else:
        completeness = ' | '.join([f"{kw} {rg}" for kw, rg in zip(keywords, ranges)])
        print('Completeness and Position are :', completeness)
        filter_start = int(ranges[0].split('-')[0])
        filter_end = int(ranges[-1].split('-')[1])
        print('specified range from %s to %s will be applied unless manually specified'%(filter_start, filter_end))
        with open('range.txt', 'w') as f:
            f.write(str(filter_start) + ' ' + str(filter_end))
        return completeness, filter_start, filter_end
