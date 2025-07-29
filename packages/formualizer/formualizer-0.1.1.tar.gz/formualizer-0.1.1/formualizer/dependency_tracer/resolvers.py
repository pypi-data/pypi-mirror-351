"""
Example resolvers for various data sources.

This module provides concrete implementations of FormulaResolver for common
data sources like openpyxl workbooks, JSON data, and in-memory dictionaries.
"""

from typing import Dict, List, Any, Optional, Iterator, Union
import json
import logging
from pathlib import Path

from formualizer import CellRef
from . import FormulaResolver

logger = logging.getLogger(__name__)


class DictResolver(FormulaResolver):
    """
    Simple resolver using nested dictionaries.

    Data structure: {sheet_name: {(row, col): {"formula": str, "value": Any}}}
    """

    def __init__(self, data: Dict[str, Dict[tuple, Dict[str, Any]]]):
        """
        Initialize with dictionary data.

        Args:
            data: Nested dict structure with sheet -> (row, col) -> cell data
        """
        self.data = data

    def get_formula(self, address: CellRef) -> Optional[str]:
        """Get formula for cell."""
        # Try both integer and string column representations
        col_int = (
            address.col
            if isinstance(address.col, int)
            else self._col_letter_to_num(address.col)
        )
        col_str = (
            address.col
            if isinstance(address.col, str)
            else self._col_num_to_letter(address.col)
        )

        sheet_data = self.data.get(address.sheet, {})
        cell_data = sheet_data.get((address.row, col_str), {}) or sheet_data.get(
            (address.row, col_int), {}
        )
        return cell_data.get("formula")

    def get_value(self, address: CellRef) -> Any:
        """Get computed value for cell."""
        # Try both integer and string column representations
        col_int = (
            address.col
            if isinstance(address.col, int)
            else self._col_letter_to_num(address.col)
        )
        col_str = (
            address.col
            if isinstance(address.col, str)
            else self._col_num_to_letter(address.col)
        )

        sheet_data = self.data.get(address.sheet, {})
        cell_data = sheet_data.get((address.row, col_str), {}) or sheet_data.get(
            (address.row, col_int), {}
        )
        return cell_data.get("value")

    def get_all_formula_cells(self) -> Iterator[CellRef]:
        """Iterate over all cells with formulas."""
        for sheet_name, sheet_data in self.data.items():
            for (row, col), cell_data in sheet_data.items():
                if "formula" in cell_data and cell_data["formula"]:
                    # Convert numeric col to letter if needed
                    col_str = (
                        col if isinstance(col, str) else self._col_num_to_letter(col)
                    )
                    yield CellRef(sheet=sheet_name, row=row, col=col_str)

    def get_sheet_names(self) -> List[str]:
        """Get available sheet names."""
        return list(self.data.keys())

    def cell_exists(self, address: CellRef) -> bool:
        """Check if cell exists."""
        # Try both integer and string column representations
        col_int = (
            address.col
            if isinstance(address.col, int)
            else self._col_letter_to_num(address.col)
        )
        col_str = (
            address.col
            if isinstance(address.col, str)
            else self._col_num_to_letter(address.col)
        )

        sheet_data = self.data.get(address.sheet, {})
        return (address.row, col_str) in sheet_data or (
            address.row,
            col_int,
        ) in sheet_data

    @staticmethod
    def _col_num_to_letter(col_num: int) -> str:
        """Convert column number to letter (1=A, 2=B, etc.)."""
        result = ""
        while col_num > 0:
            col_num -= 1
            result = chr(65 + col_num % 26) + result
            col_num //= 26
        return result

    @staticmethod
    def _col_letter_to_num(col_letter: str) -> int:
        """Convert column letter to number (A=1, B=2, etc.)."""
        num = 0
        for char in col_letter.upper():
            num = num * 26 + (ord(char) - ord("A") + 1)
        return num


class JsonResolver(FormulaResolver):
    """
    Resolver for JSON-serialized spreadsheet data.

    JSON structure:
    {
        "sheets": {
            "Sheet1": {
                "cells": {
                    "A1": {"formula": "=B1+1", "value": 42},
                    "B1": {"value": 41}
                }
            }
        }
    }
    """

    def __init__(self, json_data: Union[str, Path, Dict]):
        """
        Initialize with JSON data.

        Args:
            json_data: JSON string, file path, or parsed dict
        """
        if isinstance(json_data, (str, Path)):
            with open(json_data, "r") as f:
                self.data = json.load(f)
        else:
            self.data = json_data

        self.sheets = self.data.get("sheets", {})

    def get_formula(self, address: CellRef) -> Optional[str]:
        """Get formula for cell."""
        # Convert column to letter if it's an integer
        col_str = (
            address.col
            if isinstance(address.col, str)
            else self._col_num_to_letter(address.col)
        )
        cell_ref = f"{col_str}{address.row}"
        cell_data = (
            self.sheets.get(address.sheet, {}).get("cells", {}).get(cell_ref, {})
        )
        return cell_data.get("formula")

    def get_value(self, address: CellRef) -> Any:
        """Get computed value for cell."""
        # Convert column to letter if it's an integer
        col_str = (
            address.col
            if isinstance(address.col, str)
            else self._col_num_to_letter(address.col)
        )
        cell_ref = f"{col_str}{address.row}"
        cell_data = (
            self.sheets.get(address.sheet, {}).get("cells", {}).get(cell_ref, {})
        )
        return cell_data.get("value")

    def get_all_formula_cells(self) -> Iterator[CellRef]:
        """Iterate over all cells with formulas."""
        for sheet_name, sheet_data in self.sheets.items():
            cells = sheet_data.get("cells", {})
            for cell_ref, cell_data in cells.items():
                if "formula" in cell_data and cell_data["formula"]:
                    # Parse A1 reference
                    col = "".join(c for c in cell_ref if c.isalpha())
                    row = int("".join(c for c in cell_ref if c.isdigit()))
                    yield CellRef(sheet=sheet_name, row=row, col=col)

    def get_sheet_names(self) -> List[str]:
        """Get available sheet names."""
        return list(self.sheets.keys())

    def cell_exists(self, address: CellRef) -> bool:
        """Check if cell exists."""
        # Convert column to letter if it's an integer
        col_str = (
            address.col
            if isinstance(address.col, str)
            else self._col_num_to_letter(address.col)
        )
        cell_ref = f"{col_str}{address.row}"
        return cell_ref in self.sheets.get(address.sheet, {}).get("cells", {})

    @staticmethod
    def _col_num_to_letter(col_num: int) -> str:
        """Convert column number to letter (1=A, 2=B, etc.)."""
        result = ""
        while col_num > 0:
            col_num -= 1
            result = chr(65 + col_num % 26) + result
            col_num //= 26
        return result


class OpenpyxlResolver(FormulaResolver):
    """
    Resolver for openpyxl workbooks.

    Requires openpyxl to be installed: pip install openpyxl
    """

    def __init__(self, workbook_path: Union[str, Path] = None, workbook=None):
        """
        Initialize with openpyxl workbook.

        Args:
            workbook_path: Path to Excel file
            workbook: Existing openpyxl workbook object
        """
        try:
            import openpyxl

            self.openpyxl = openpyxl
        except ImportError:
            raise ImportError(
                "openpyxl is required for OpenpyxlResolver. Install with: pip install openpyxl"
            )

        if workbook is not None:
            self.workbook = workbook
        elif workbook_path is not None:
            self.workbook = openpyxl.load_workbook(workbook_path, data_only=False)
        else:
            raise ValueError("Either workbook_path or workbook must be provided")

    def get_formula(self, address: CellRef) -> Optional[str]:
        """Get formula for cell."""
        try:
            worksheet = self.workbook[address.sheet]
            cell = worksheet[f"{address.str_col}{address.row}"]

            # openpyxl stores formulas without the leading =
            if cell.data_type == "f" and cell.value:
                return cell.value
            return None
        except (KeyError, AttributeError):
            return None

    def get_value(self, address: CellRef) -> Any:
        """Get computed value for cell."""
        try:
            # Load workbook with data_only=True for calculated values
            if not hasattr(self, "_data_workbook"):
                self._data_workbook = self.openpyxl.load_workbook(
                    self.workbook.path if hasattr(self.workbook, "path") else None,
                    data_only=True,
                )

            worksheet = self._data_workbook[address.sheet]
            cell = worksheet[f"{address.col}{address.row}"]
            return cell.value
        except (KeyError, AttributeError):
            return None

    def get_all_formula_cells(self) -> Iterator[CellRef]:
        """Iterate over all cells with formulas."""
        for sheet_name in self.workbook.sheetnames:
            worksheet = self.workbook[sheet_name]
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.data_type == "f" and cell.value:
                        # Convert to column letter
                        col_letter = self.openpyxl.utils.get_column_letter(cell.column)
                        yield CellRef(sheet=sheet_name, row=cell.row, col=col_letter)

    def get_sheet_names(self) -> List[str]:
        """Get available sheet names."""
        return self.workbook.sheetnames

    def cell_exists(self, address: CellRef) -> bool:
        """Check if cell exists and has content."""
        try:
            worksheet = self.workbook[address.sheet]
            cell = worksheet[f"{address.col}{address.row}"]
            return cell.value is not None
        except (KeyError, AttributeError):
            return False


class FastexcelResolver(FormulaResolver):
    """
    Resolver for fastexcel (calamine-based) workbooks.

    Requires fastexcel to be installed: pip install fastexcel
    Note: fastexcel is read-only and may have different formula handling
    """

    def __init__(self, workbook_path: Union[str, Path]):
        """
        Initialize with fastexcel workbook.

        Args:
            workbook_path: Path to Excel file
        """
        try:
            import fastexcel

            self.fastexcel = fastexcel
        except ImportError:
            raise ImportError(
                "fastexcel is required for FastexcelResolver. Install with: pip install fastexcel"
            )

        self.workbook_path = Path(workbook_path)
        self.workbook = fastexcel.read_excel(self.workbook_path)

        # Cache sheet data
        self._sheet_cache = {}
        for sheet_name in self.get_sheet_names():
            try:
                self._sheet_cache[sheet_name] = self.workbook[sheet_name].to_pandas()
            except Exception as e:
                logger.warning(f"Could not load sheet {sheet_name}: {e}")
                self._sheet_cache[sheet_name] = None

    def get_formula(self, address: CellRef) -> Optional[str]:
        """
        Get formula for cell.

        Note: fastexcel may not preserve formulas - this is a limitation
        """
        # fastexcel typically doesn't preserve formulas, only calculated values
        logger.warning(
            "FastexcelResolver: Formula extraction not supported, returning None"
        )
        return None

    def get_value(self, address: CellRef) -> Any:
        """Get computed value for cell."""
        try:
            df = self._sheet_cache.get(address.sheet)
            if df is None:
                return None

            # Convert column letter to index
            col_idx = self._col_letter_to_num(address.col) - 1
            row_idx = address.row - 1  # Convert to 0-based

            if row_idx < len(df) and col_idx < len(df.columns):
                return df.iloc[row_idx, col_idx]
            return None
        except (KeyError, IndexError, AttributeError):
            return None

    def get_all_formula_cells(self) -> Iterator[CellRef]:
        """
        Iterate over all cells with formulas.

        Note: Since fastexcel doesn't preserve formulas, this returns empty
        """
        logger.warning("FastexcelResolver: Cannot identify formula cells")
        return iter([])

    def get_sheet_names(self) -> List[str]:
        """Get available sheet names."""
        return list(self.workbook.sheet_names)

    def cell_exists(self, address: CellRef) -> bool:
        """Check if cell exists and has content."""
        try:
            value = self.get_value(address)
            return value is not None and str(value).strip() != ""
        except:
            return False

    @staticmethod
    def _col_letter_to_num(col_letter: str) -> int:
        """Convert column letter to number (A=1, B=2, etc.)."""
        num = 0
        for char in col_letter.upper():
            num = num * 26 + (ord(char) - ord("A") + 1)
        return num


class CombinedResolver(FormulaResolver):
    """
    Resolver that combines multiple resolvers with priority order.

    Useful for scenarios where you want to overlay data from multiple sources,
    e.g., JSON metadata with openpyxl for actual values.
    """

    def __init__(self, resolvers: List[FormulaResolver]):
        """
        Initialize with list of resolvers in priority order.

        Args:
            resolvers: List of resolvers, first has highest priority
        """
        if not resolvers:
            raise ValueError("At least one resolver must be provided")
        self.resolvers = resolvers

    def get_formula(self, address: CellRef) -> Optional[str]:
        """Get formula from first resolver that has it."""
        for resolver in self.resolvers:
            formula = resolver.get_formula(address)
            if formula is not None:
                return formula
        return None

    def get_value(self, address: CellRef) -> Any:
        """Get value from first resolver that has it."""
        for resolver in self.resolvers:
            try:
                value = resolver.get_value(address)
                if value is not None:
                    return value
            except Exception:
                continue
        return None

    def get_all_formula_cells(self) -> Iterator[CellRef]:
        """Get formula cells from all resolvers (deduplicated)."""
        seen = set()
        for resolver in self.resolvers:
            for address in resolver.get_all_formula_cells():
                if address not in seen:
                    seen.add(address)
                    yield address

    def get_sheet_names(self) -> List[str]:
        """Get sheet names from all resolvers (deduplicated)."""
        sheet_names = set()
        for resolver in self.resolvers:
            sheet_names.update(resolver.get_sheet_names())
        return sorted(sheet_names)

    def cell_exists(self, address: CellRef) -> bool:
        """Check if cell exists in any resolver."""
        return any(resolver.cell_exists(address) for resolver in self.resolvers)
