"""
Practical examples demonstrating dependency tracing with real data sources.

This module shows how to use the dependency tracer with various data sources
including openpyxl, JSON, and custom resolvers.
"""

import json
from pathlib import Path
from typing import Dict, Any, Optional
import tempfile

from . import DependencyTracer
from .resolvers import JsonResolver, DictResolver, CombinedResolver

try:
    from .resolvers import OpenpyxlResolver

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def example_1_simple_json_tracing():
    """
    Example 1: Basic dependency tracing with JSON data.

    Shows how to trace precedents and dependents in a simple financial model.
    """
    print("=== Example 1: Simple JSON Dependency Tracing ===\n")

    # Create a simple financial model
    financial_model = {
        "sheets": {
            "Inputs": {
                "cells": {
                    "B2": {"value": 1000, "note": "Initial Investment"},
                    "B3": {"value": 0.05, "note": "Interest Rate"},
                    "B4": {"value": 10, "note": "Years"},
                }
            },
            "Calculations": {
                "cells": {
                    "B2": {"formula": "=Inputs!B2", "value": 1000, "note": "Principal"},
                    "B3": {"formula": "=Inputs!B3", "value": 0.05, "note": "Rate"},
                    "B4": {"formula": "=Inputs!B4", "value": 10, "note": "Time"},
                    "B6": {
                        "formula": "=B2*(1+B3)^B4",
                        "value": 1628.89,
                        "note": "Future Value",
                    },
                    "B7": {
                        "formula": "=B6-B2",
                        "value": 628.89,
                        "note": "Interest Earned",
                    },
                }
            },
            "Summary": {
                "cells": {
                    "B2": {
                        "formula": "=Calculations!B6",
                        "value": 1628.89,
                        "note": "Total Value",
                    },
                    "B3": {
                        "formula": "=Calculations!B7",
                        "value": 628.89,
                        "note": "Total Interest",
                    },
                    "B4": {
                        "formula": "=B3/Calculations!B2*100",
                        "value": 62.89,
                        "note": "ROI %",
                    },
                }
            },
        }
    }

    # Set up tracer
    resolver = JsonResolver(financial_model)
    tracer = DependencyTracer(resolver)

    print("📊 Financial Model Dependency Analysis")
    print(f"📋 Found {len(list(resolver.get_all_formula_cells()))} formula cells\n")

    # Example: What affects the final ROI calculation?
    roi_cell = "Summary!B4"
    print(f"🔍 What affects {roi_cell} (ROI %)?")
    precedents = tracer.trace_precedents(roi_cell, recursive=True)

    print("   Direct and indirect precedents:")
    for p in sorted(precedents, key=lambda x: (x.sheet, x.col, x.row)):
        formula = resolver.get_formula(p)
        value = resolver.get_value(p)
        print(f"   - {p}: {formula or f'={value}'}")

    print(f"\n🎯 What would be affected if we change the interest rate (Inputs!B3)?")
    rate_dependents = tracer.trace_dependents("Inputs!B3", recursive=True)

    print("   Affected cells:")
    for d in sorted(rate_dependents, key=lambda x: (x.sheet, x.col, x.row)):
        formula = resolver.get_formula(d)
        value = resolver.get_value(d)
        print(f"   - {d}: {formula} = {value}")

    print("\n" + "=" * 60 + "\n")


def example_2_openpyxl_integration():
    """
    Example 2: Integration with openpyxl for real Excel files.

    Creates a sample Excel file and demonstrates dependency tracing.
    """
    print("=== Example 2: openpyxl Integration ===\n")

    if not OPENPYXL_AVAILABLE:
        print("⚠️  openpyxl not available. Install with: pip install openpyxl")
        print("   Skipping openpyxl example.\n")
        return

    import openpyxl
    from .resolvers import OpenpyxlResolver

    # Create a sample workbook
    wb = openpyxl.Workbook()

    # Remove default sheet and add our sheets
    wb.remove(wb.active)

    # Sales Data sheet
    sales_sheet = wb.create_sheet("Sales")
    sales_data = [
        ["Month", "Sales", "Costs"],
        ["Jan", 1000, 600],
        ["Feb", 1200, 720],
        ["Mar", 1100, 660],
        ["Apr", 1300, 780],
    ]

    for row_idx, row_data in enumerate(sales_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            sales_sheet.cell(row=row_idx, column=col_idx, value=value)

    # Summary sheet with formulas
    summary_sheet = wb.create_sheet("Summary")
    summary_sheet["A1"] = "Total Sales"
    summary_sheet["B1"] = "=SUM(Sales!B2:B5)"
    summary_sheet["A2"] = "Total Costs"
    summary_sheet["B2"] = "=SUM(Sales!C2:C5)"
    summary_sheet["A3"] = "Profit"
    summary_sheet["B3"] = "=B1-B2"
    summary_sheet["A4"] = "Margin %"
    summary_sheet["B4"] = "=B3/B1*100"

    # Analysis sheet
    analysis_sheet = wb.create_sheet("Analysis")
    analysis_sheet["A1"] = "Avg Monthly Sales"
    analysis_sheet["B1"] = "=Summary!B1/4"
    analysis_sheet["A2"] = "Profit per Month"
    analysis_sheet["B2"] = "=Summary!B3/4"

    # Save to temporary file
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        wb.save(tmp.name)
        temp_path = tmp.name

    try:
        # Set up dependency tracer
        resolver = OpenpyxlResolver(temp_path)
        tracer = DependencyTracer(resolver)

        print("📊 Excel Workbook Dependency Analysis")
        print(f"📋 Sheets: {', '.join(resolver.get_sheet_names())}")

        formula_cells = list(resolver.get_all_formula_cells())
        print(f"📋 Found {len(formula_cells)} formula cells\n")

        # Show all formulas
        print("📝 All formulas in workbook:")
        for cell in sorted(formula_cells, key=lambda x: (x.sheet, x.col, x.row)):
            formula = resolver.get_formula(cell)
            print(f"   {cell}: {formula}")

        print(f"\n🔍 What affects the profit margin (Summary!B4)?")
        margin_precedents = tracer.trace_precedents("Summary!B4", recursive=True)
        print("   Dependencies:")
        for p in sorted(margin_precedents, key=lambda x: (x.sheet, x.col, x.row)):
            value = resolver.get_value(p)
            print(f"   - {p}: {value}")

        print(f"\n🎯 What would change if we modify January sales (Sales!B2)?")
        jan_dependents = tracer.trace_dependents("Sales!B2", recursive=True)
        print("   Affected formulas:")
        for d in sorted(jan_dependents, key=lambda x: (x.sheet, x.col, x.row)):
            formula = resolver.get_formula(d)
            print(f"   - {d}: {formula}")

        print(f"\n📈 Evaluation order (topological sort):")
        try:
            topo_order = tracer.topological_sort()
            for i, cell in enumerate(topo_order, 1):
                formula = resolver.get_formula(cell)
                print(f"   {i:2d}. {cell}: {formula}")
        except ValueError as e:
            print(f"   Error: {e}")

    finally:
        # Clean up
        Path(temp_path).unlink()

    print("\n" + "=" * 60 + "\n")


def example_3_combined_resolvers():
    """
    Example 3: Using combined resolvers for metadata overlay.

    Shows how to combine multiple data sources with priority handling.
    """
    print("=== Example 3: Combined Resolvers ===\n")

    # Base workbook data
    base_data = {
        "sheets": {
            "Data": {
                "cells": {
                    "A1": {"formula": "=B1+C1", "value": 30},
                    "B1": {"value": 10},
                    "C1": {"value": 20},
                    "A2": {"formula": "=A1*2", "value": 60},
                }
            }
        }
    }

    # Metadata overlay (formulas with descriptions, alternative formulas, etc.)
    metadata_overlay = {
        "Data": {
            (1, "A"): {
                "formula": "=B1+C1",
                "description": "Sum of base values",
                "category": "calculation",
            },
            (2, "A"): {
                "formula": "=A1*2",
                "description": "Double the sum",
                "category": "transformation",
            },
            (1, "B"): {"description": "Base value 1", "category": "input"},
            (1, "C"): {"description": "Base value 2", "category": "input"},
        }
    }

    # Create resolvers
    base_resolver = JsonResolver(base_data)
    metadata_resolver = DictResolver(metadata_overlay)

    # Combine with metadata taking priority for formulas
    combined_resolver = CombinedResolver([metadata_resolver, base_resolver])

    # Custom resolver that adds description information
    class DescriptiveResolver:
        def __init__(self, base_resolver, metadata_resolver):
            self.base = base_resolver
            self.metadata = metadata_resolver

        def get_formula_with_description(self, address):
            formula = self.base.get_formula(address)
            if formula:
                # Try to get description from metadata
                try:
                    desc_data = self.metadata.data.get(address.sheet, {}).get(
                        (address.row, address.col), {}
                    )
                    description = desc_data.get("description", "")
                    category = desc_data.get("category", "unknown")
                    return {
                        "formula": formula,
                        "description": description,
                        "category": category,
                        "value": self.base.get_value(address),
                    }
                except:
                    return {"formula": formula, "value": self.base.get_value(address)}
            return None

    descriptive_resolver = DescriptiveResolver(base_resolver, metadata_resolver)
    tracer = DependencyTracer(combined_resolver)

    print("📊 Combined Resolver Analysis")
    print("🔍 Enhanced formula information:\n")

    formula_cells = list(combined_resolver.get_all_formula_cells())
    for cell in sorted(formula_cells, key=lambda x: (x.sheet, x.col, x.row)):
        info = descriptive_resolver.get_formula_with_description(cell)
        if info:
            print(f"   {cell}:")
            print(f"      Formula: {info['formula']}")
            print(f"      Value: {info['value']}")
            print(f"      Description: {info.get('description', 'N/A')}")
            print(f"      Category: {info.get('category', 'N/A')}")
            print()

    print("🎯 Dependency analysis with categories:")
    precedents = tracer.trace_precedents("Data!A2", recursive=True)

    for p in sorted(precedents, key=lambda x: (x.sheet, x.col, x.row)):
        info = descriptive_resolver.get_formula_with_description(p)
        category = info.get("category", "unknown") if info else "data"
        formula = combined_resolver.get_formula(p)
        value = combined_resolver.get_value(p)

        print(f"   {p} [{category}]: {formula or f'={value}'}")

    print("\n" + "=" * 60 + "\n")


def example_4_cycle_detection():
    """
    Example 4: Circular dependency detection and handling.

    Shows how the tracer handles circular references and provides debugging info.
    """
    print("=== Example 4: Circular Dependency Detection ===\n")

    # Create data with circular references
    circular_data = {
        "sheets": {
            "Sheet1": {
                "cells": {
                    # Simple circular reference
                    "A1": {"formula": "=B1+1", "value": "#REF!"},
                    "B1": {"formula": "=A1+1", "value": "#REF!"},
                    # Longer cycle
                    "C1": {"formula": "=D1*2", "value": "#REF!"},
                    "D1": {"formula": "=E1+5", "value": "#REF!"},
                    "E1": {"formula": "=C1/3", "value": "#REF!"},
                    # Valid formulas that don't participate in cycles
                    "F1": {"formula": "=10", "value": 10},
                    "G1": {"formula": "=F1*2", "value": 20},
                    "H1": {"formula": "=G1+F1", "value": 30},
                }
            }
        }
    }

    resolver = JsonResolver(circular_data)
    tracer = DependencyTracer(resolver)

    print("📊 Circular Dependency Analysis")

    # Find all circular dependencies
    cycles = tracer.find_circular_dependencies()

    if cycles:
        print(f"⚠️  Found {len(cycles)} circular dependency chain(s):\n")
        for i, cycle in enumerate(cycles, 1):
            print(f"   Cycle {i}: {' → '.join(str(cell) for cell in cycle)}")

            # Show the formulas in the cycle
            print("   Formulas:")
            for cell in cycle[:-1]:  # Exclude duplicate last cell
                formula = resolver.get_formula(cell)
                print(f"      {cell}: {formula}")
            print()
    else:
        print("✅ No circular dependencies found")

    # Try topological sort
    print("📈 Attempting topological sort:")
    try:
        topo_order = tracer.topological_sort()
        print("✅ Successfully sorted (no cycles in dependency graph)")
        print(f"   Evaluation order for {len(topo_order)} cells:")
        for cell in topo_order:
            formula = resolver.get_formula(cell)
            print(f"      {cell}: {formula}")
    except ValueError as e:
        print(f"❌ {e}")

    print(f"\n🔍 Manual precedent tracing (with max depth protection):")

    # Show how max_depth prevents infinite recursion
    for cell_ref in ["Sheet1!A1", "Sheet1!C1"]:
        print(f"\n   Tracing {cell_ref} with max_depth=3:")
        try:
            precedents = tracer.trace_precedents(cell_ref, max_depth=3)
            print(
                f"      Found {len(precedents)} precedents: {sorted(precedents, key=str)}"
            )
        except Exception as e:
            print(f"      Error: {e}")

    print("\n" + "=" * 60 + "\n")


def example_5_performance_and_caching():
    """
    Example 5: Performance optimization and cache management.

    Demonstrates caching behavior and performance considerations.
    """
    print("=== Example 5: Performance and Caching ===\n")

    # Create a larger dataset to demonstrate caching
    large_data = {"sheets": {"Sheet1": {"cells": {}}}}

    # Generate a chain of dependencies: A1->A2->A3->...->A100
    for i in range(1, 101):
        if i == 1:
            large_data["sheets"]["Sheet1"]["cells"][f"A{i}"] = {
                "formula": "=10",
                "value": 10,
            }
        else:
            large_data["sheets"]["Sheet1"]["cells"][f"A{i}"] = {
                "formula": f"=A{i - 1}+1",
                "value": 10 + i - 1,
            }

    # Add some branching dependencies
    for i in range(1, 51):
        large_data["sheets"]["Sheet1"]["cells"][f"B{i}"] = {
            "formula": f"=A{i}*2",
            "value": (10 + i - 1) * 2,
        }

    resolver = JsonResolver(large_data)

    print("📊 Performance Analysis")
    print(f"📋 Dataset: {len(large_data['sheets']['Sheet1']['cells'])} cells")

    # Test with caching enabled
    import time

    print("\n🚀 Testing with caching enabled:")
    tracer_cached = DependencyTracer(resolver, enable_caching=True)

    start_time = time.time()
    precedents_a100 = tracer_cached.trace_precedents("Sheet1!A100", recursive=True)
    cached_time = time.time() - start_time

    print(
        f"   First trace (A100 precedents): {len(precedents_a100)} precedents in {cached_time:.4f}s"
    )

    # Trace again (should be faster due to caching)
    start_time = time.time()
    precedents_a100_again = tracer_cached.trace_precedents(
        "Sheet1!A100", recursive=True
    )
    cached_time_2 = time.time() - start_time

    print(
        f"   Second trace (cached): {len(precedents_a100_again)} precedents in {cached_time_2:.4f}s"
    )
    print(f"   Speedup: {cached_time / cached_time_2:.1f}x faster")

    # Show cache statistics
    cache_stats = tracer_cached.get_stats()
    print(f"\n📈 Cache Statistics:")
    for key, value in cache_stats.items():
        print(f"   {key}: {value}")

    # Test cache invalidation
    print(f"\n🔄 Testing cache invalidation:")
    print(
        f"   Cache size before invalidation: {cache_stats.get('dependency_cache_size', 0)}"
    )

    tracer_cached.invalidate_cell("Sheet1!A50")

    new_stats = tracer_cached.get_stats()
    print(
        f"   Cache size after invalidating A50: {new_stats.get('dependency_cache_size', 0)}"
    )

    # Test without caching
    print(f"\n🐌 Testing without caching:")
    tracer_uncached = DependencyTracer(resolver, enable_caching=False)

    start_time = time.time()
    precedents_uncached = tracer_uncached.trace_precedents(
        "Sheet1!A100", recursive=True
    )
    uncached_time = time.time() - start_time

    print(
        f"   Trace without caching: {len(precedents_uncached)} precedents in {uncached_time:.4f}s"
    )

    if cached_time_2 > 0:
        print(
            f"   Caching advantage: {uncached_time / cached_time_2:.1f}x faster with cache"
        )

    # Show memory usage considerations
    print(f"\n💾 Memory Considerations:")
    print(f"   Formula cells: {len(list(resolver.get_all_formula_cells()))}")
    print(f"   Cache entries: {cache_stats.get('dependency_cache_size', 0)}")
    print(f"   AST cache: {cache_stats.get('ast_cache_size', 0)}")

    if hasattr(tracer_cached.resolver, "cache_stats"):
        resolver_stats = tracer_cached.resolver.cache_stats()
        print(f"   Resolver cache: {resolver_stats}")

    print("\n" + "=" * 60 + "\n")


def run_all_examples():
    """Run all examples in sequence."""
    print("🚀 Formualizer Dependency Tracer Examples\n")
    print("This demonstrates a robust, resolver-agnostic dependency tracing system")
    print("with intelligent caching and support for various data sources.\n")

    examples = [
        example_1_simple_json_tracing,
        example_2_openpyxl_integration,
        example_3_combined_resolvers,
        example_4_cycle_detection,
        example_5_performance_and_caching,
    ]

    for example_func in examples:
        try:
            example_func()
        except Exception as e:
            print(f"❌ Error in {example_func.__name__}: {e}\n")
            continue

    print("✅ All examples completed!")
    print("\nKey Features Demonstrated:")
    print("• ABC-based resolver architecture for flexible data source integration")
    print("• Intelligent caching with selective invalidation")
    print("• Precedent and dependent tracing with cycle detection")
    print("• Cross-sheet and recursive dependency analysis")
    print("• Performance optimization and memory management")
    print("• Integration with openpyxl, JSON, and custom data sources")


if __name__ == "__main__":
    run_all_examples()
