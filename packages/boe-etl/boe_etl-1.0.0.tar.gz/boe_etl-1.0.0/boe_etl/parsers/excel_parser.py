"""
Excel file parsing functionality for the ETL pipeline.
"""
import logging
from pathlib import Path
from typing import Dict, Any, List, Union, Optional
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter

from .base_parser import BaseParser

logger = logging.getLogger(__name__)


class ExcelParser(BaseParser):
    """Parser for Excel files in the ETL pipeline."""
    
    def __init__(self, bank: str, quarter: str):
        """Initialize the Excel parser.
        
        Args:
            bank: Name of the bank
            quarter: Quarter identifier (e.g., 'Q1_2025')
        """
        super().__init__(bank, quarter)
        self.supported_formats = {'.xlsx', '.xls'}
    
    def parse(self, file_path: Path) -> Dict[str, Any]:
        """Parse an Excel file and extract its content.
        
        Args:
            file_path: Path to the Excel file
            
        Returns:
            Dictionary containing the parsed data with standardized structure
        """
        if not file_path.exists():
            raise FileNotFoundError(f"Excel file not found: {file_path}")
            
        # Infer document type from filename
        document_type = self._infer_document_type(file_path.name)
        
        result = {
            'bank': self.bank,
            'quarter': self.quarter,
            'file_path': str(file_path),
            'document_type': document_type,
            'content': {
                'tables': {},
                'metadata': {}
            },
            'metadata': self.get_metadata(file_path)
        }
        
        try:
            # Read the Excel file
            xls = pd.ExcelFile(file_path)
            result['content']['metadata']['sheet_names'] = xls.sheet_names
            
            # Parse each sheet
            for sheet_name in xls.sheet_names:
                try:
                    # Read the sheet into a DataFrame
                    df = pd.read_excel(file_path, sheet_name=sheet_name)
                    
                    # Convert DataFrame to a list of dictionaries (one per row)
                    records = df.where(pd.notnull(df), None).to_dict('records')
                    
                    # Store the sheet data
                    result['content']['tables'][sheet_name] = {
                        'columns': [{'name': col, 'dtype': str(dtype)} 
                                  for col, dtype in df.dtypes.items()],
                        'data': records,
                        'shape': {
                            'rows': df.shape[0],
                            'columns': df.shape[1]
                        }
                    }
                    
                    # Add table-level metadata
                    self._extract_table_metadata(
                        file_path, 
                        sheet_name, 
                        result['content']['tables'][sheet_name]
                    )
                    
                except Exception as e:
                    logger.warning(f"Error parsing sheet '{sheet_name}': {str(e)}")
                    result['content']['tables'][sheet_name] = {
                        'error': str(e),
                        'error_type': type(e).__name__
                    }
            
            # Add workbook-level metadata
            self._extract_workbook_metadata(file_path, result['content']['metadata'])
            
            return result
            
        except Exception as e:
            logger.error(f"Error parsing Excel file {file_path}: {str(e)}")
            raise
    
    def _infer_document_type(self, filename: str) -> str:
        """Infer the document type from the filename.
        
        Args:
            filename: Name of the file
            
        Returns:
            Inferred document type
        """
        filename = filename.lower()
        if any(term in filename for term in ['result', 'rslt']):
            return 'results'
        elif any(term in filename for term in ['supplement', 'fsqtr']):
            return 'supplement'
        elif 'financial' in filename:
            return 'financial_statements'
        elif 'presentation' in filename:
            return 'presentation'
        return 'other'
    
    def _extract_table_metadata(self, file_path: Path, sheet_name: str, table_data: Dict):
        """Extract metadata specific to an Excel table.
        
        Args:
            file_path: Path to the Excel file
            sheet_name: Name of the sheet
            table_data: Dictionary to store the extracted metadata
        """
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            ws = wb[sheet_name]
            
            # Add sheet-level metadata
            table_data['metadata'] = {
                'has_formulas': self._has_formulas(ws),
                'has_merged_cells': bool(ws.merged_cells.ranges),
                'has_charts': len(ws._charts) > 0,
                'has_images': len(ws._images) > 0,
                'dimensions': ws.calculate_dimension(),
                'freeze_panes': str(ws.freeze_panes) if ws.freeze_panes else None
            }
            
            # Add data validation info
            if ws.data_validations:
                table_data['metadata']['data_validations'] = [
                    str(dv) for dv in ws.data_validations.dataValidation
                ]
                
        except Exception as e:
            logger.warning(f"Error extracting metadata for sheet '{sheet_name}': {str(e)}")
            table_data['metadata'] = {'error': str(e)}
    
    def _extract_workbook_metadata(self, file_path: Path, metadata: Dict):
        """Extract workbook-level metadata.
        
        Args:
            file_path: Path to the Excel file
            metadata: Dictionary to store the metadata
        """
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True)
            
            # Add document properties
            props = wb.properties
            metadata.update({
                'creator': props.creator,
                'last_modified_by': props.lastModifiedBy,
                'created': props.created.isoformat() if props.created else None,
                'modified': props.modified.isoformat() if props.modified else None,
                'has_macros': 'vba' in wb.vba_archive.namelist() if wb.vba_archive else False,
                'defined_names': [str(name) for name in wb.defined_names.definedName],
                'active_sheet': wb.active.title if wb.active else None,
                'sheet_count': len(wb.sheetnames)
            })
            
        except Exception as e:
            logger.warning(f"Error extracting workbook metadata: {str(e)}")
    
    @staticmethod
    def _has_formulas(worksheet) -> bool:
        """Check if a worksheet contains any formulas."""
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.data_type == 'f':  # 'f' indicates a formula
                    return True
        return False


def parse_excel(bank: str, quarter: str, file_path: Path, document_type: str = "unknown") -> Dict[str, Any]:
    """
    Parse an Excel file and extract its content.
    
    This is a convenience function that creates an ExcelParser instance.
    For new code, it's recommended to use the ExcelParser class directly.
    
    Args:
        bank: Name of the bank
        quarter: Quarter identifier (e.g., 'Q1_2025')
        file_path: Path to the Excel file
        document_type: Type of document (e.g., 'results', 'supplement', 'financial')
                      If "unknown", will attempt to infer from filename
    
    Returns:
        Dictionary containing the parsed data
    """
    parser = ExcelParser(bank, quarter)
    result = parser.parse(file_path)
    
    # For backward compatibility
    if document_type != "unknown" and document_type != result['document_type']:
        result['document_type'] = document_type
        
    return result


def test_parse_excel():
    """Test function for the Excel parser"""
    import tempfile
    import pandas as pd
    from pathlib import Path
    
    # Create a simple Excel file for testing
    df1 = pd.DataFrame({'A': [1, 2, 3], 'B': ['a', 'b', 'c']})
    df2 = pd.DataFrame({'X': [1.1, 2.2], 'Y': ['test', 'data']})
    
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
        with pd.ExcelWriter(tmp.name, engine='openpyxl') as writer:
            df1.to_excel(writer, sheet_name='Sheet1', index=False)
            df2.to_excel(writer, sheet_name='Sheet2', index=False)
        
        test_file = Path(tmp.name)
    
    try:
        parser = ExcelParser('test_bank', 'Q1_2025')
        result = parser.parse(test_file)
        print("Excel parse test successful!")
        print(f"Sheets: {list(result['content']['tables'].keys())}")
        print(f"Document type: {result['document_type']}")
        return True
    except Exception as e:
        print(f"Excel parse test failed: {str(e)}")
        return False
    finally:
        test_file.unlink()  # Clean up


if __name__ == "__main__":
    test_parse_excel()
