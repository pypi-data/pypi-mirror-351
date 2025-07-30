import matplotlib.pyplot as plt
import numpy as np
import pickle
from openpyxl import Workbook, load_workbook
import pandas as pd
from typing import Union


def show_boxplot(data: Union[list, np.ndarray], xlabels: str, ylabel: str, filename: Union[str, None] = None,
                 colors: Union[str, list, None] = None):
    """
    Show the boxplot of numerical data.

    Args:
        data: Numerical dataset(s).
        xlabels: Labels for each dataset. Length must be compatible with dimensions of `data`.
        ylabel: Label of y-axis.
        filename: Filename of the figure to be saved. If None, the figure is not saved.
        colors: Colors for each boxplot. If None, random colors are generated for each boxplot.
    """
    plt.rcParams['font.size'] = '10'
    plt.figure(figsize=(12, 10))
    plt.ylabel(ylabel)
    basic = dict(linestyle='-', linewidth=1.5, color='black')
    mean = dict(linestyle='--', linewidth=1.5, color='black')
    flier = dict(marker='+', linewidth=1.5, markersize=15)
    bp = plt.boxplot(data, patch_artist=True, showmeans=True, labels=xlabels, meanprops=mean, meanline=True,
                     medianprops=basic, boxprops=basic, whiskerprops=basic, flierprops=flier, capprops=basic,
                     widths=0.3, manage_ticks=True)
    plt.legend([bp['medians'][0], bp['means'][0]], ['Median', 'Mean'], loc='best')

    if colors is not None:
        for patch, color in zip(bp['boxes'], colors):
            patch.set(facecolor=color)
    else:
        rs = np.random.randint(0, 255, len(xlabels))
        gs = np.random.randint(0, 255, len(xlabels))
        bs = np.random.randint(0, 255, len(xlabels))
        colors = [f"#{r:02x}{g:02x}{b:02x}" for r, g, b in zip(rs, gs, bs)]
        for patch, color in zip(bp['boxes'], colors):
            patch.set(facecolor=color)
    plt.show()
    if filename is not None:
        plt.savefig(filename)
    return bp


def write_excel_file(data: dict, filename: str, sheet_title: str, create_new_sheet: bool = False):
    """
    Write an Excel file.
    """
    wb = Workbook()
    if create_new_sheet:
        ws= wb.create_sheet(sheet_title)
    else:
        ws = wb.active
        ws.title = sheet_title
    ws.append(list(data.keys()))
    for row in zip(*data.values()):
        ws.append(row)
    wb.save(filename)
    wb.close()


def read_excel_file(filename: str, sheet_title: Union[str, None] = None):
    """
    Read an Excel file.
    """
    wb = load_workbook(filename=filename)
    if sheet_title is not None:
        ws = wb[sheet_title]
    else:
        ws = wb.active
    data = {}
    for n, col in enumerate(ws.iter_cols(values_only=True)):
        data[f'Column #{n}'] = col
    return data


def write_csv_file(data: dict, filename: str):
    """
    Write a csv file.
    """
    df = pd.DataFrame(data)
    df.to_csv(filename, index=True)


def read_csv_file(filename: str):
    """
    Read a csv file.
    """
    data = pd.read_csv(filename)
    return data


def write_pkl_file(data: dict, filename: str):
    """
    Write a pickle file.
    """
    with open(filename, 'wb') as f:
        pickle.dump(data, f, pickle.HIGHEST_PROTOCOL)


def read_pkl_file(filename: str):
    """
    Read a pickle file.
    """
    with open(filename, 'rb') as f:
        data = pickle.load(f)
    return data
